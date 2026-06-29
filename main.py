import sys
import sqlite3
import os
import re
import csv
import ast
import uuid
import shutil
import datetime
import traceback
import qtawesome as qta
from contextlib import redirect_stdout, redirect_stderr

try:
    import pandas as pd
except ImportError:
    pd = None

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, 
    QTreeWidget, QTreeWidgetItem, QPushButton, QLineEdit, QLabel, QTableView, 
    QHeaderView, QMessageBox, QFileDialog, QListWidget, QListWidgetItem, QDialog,
    QSplitter, QScrollArea, QComboBox, QInputDialog, QFormLayout, QSpinBox,
    QDoubleSpinBox, QTextEdit, QDateTimeEdit, QCheckBox, QPlainTextEdit, QTabWidget,
    QMenuBar, QMenu, QRadioButton, QGroupBox, QDialogButtonBox, QTextBrowser
)
from PySide6.QtCore import Qt, QSettings, QDateTime, QRegularExpression, QUrl
from PySide6.QtSql import QSqlDatabase, QSqlQueryModel, QSqlQuery
from PySide6.QtGui import QAction, QFontDatabase, QFont, QRegularExpressionValidator, QDesktopServices

from class_builder_dialog import ClassBuilderDialog, DiscreteTypeBuilderDialog, init_db, get_app_icon, sanitize_name, qid

# ==========================================
# FILE ATTACHMENTS
# ==========================================
# A "file" attribute stores a reference string of the form  <uuid32hex>__<original name>
# which is ALSO the file's name on disk inside the per-database files folder.
# The generated views expose only the display part (substr from char 35), so the table,
# look-throughs and relationship titles all show the clean filename automatically.
FILES_SUBDIR_SUFFIX = "_files"
FILE_PREFIX_LEN = 34  # uuid4().hex (32) + "__"


def sanitize_filename(name):
    name = os.path.basename(str(name)).strip()
    name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '_', name)
    return name or "file"


def make_stored_filename(original):
    """Build a collision-proof on-disk/reference name that keeps the original for display."""
    return f"{uuid.uuid4().hex}__{sanitize_filename(original)}"


def display_file_name(stored_value):
    if not stored_value:
        return ""
    return str(stored_value)[FILE_PREFIX_LEN:]


def files_dir_for(db_path, create=False):
    if not db_path:
        return None
    base = os.path.dirname(os.path.abspath(db_path))
    stem = os.path.splitext(os.path.basename(db_path))[0]
    d = os.path.join(base, stem + FILES_SUBDIR_SUFFIX)
    if create:
        os.makedirs(d, exist_ok=True)
    return d


def resolve_file_path(db_path, stored_value):
    d = files_dir_for(db_path, create=False)
    if not d or not stored_value:
        return None
    return os.path.join(d, str(stored_value))


def trash_stored_file(db_path, stored_value):
    """Move a stored file into the _trash folder (best-effort, recoverable)."""
    if not stored_value:
        return
    src = resolve_file_path(db_path, stored_value)
    if not src or not os.path.exists(src):
        return
    try:
        trash = os.path.join(files_dir_for(db_path, create=True), "_trash")
        os.makedirs(trash, exist_ok=True)
        dest = os.path.join(trash, str(stored_value))
        if os.path.exists(dest):
            dest = os.path.join(trash, f"{uuid.uuid4().hex[:8]}_{stored_value}")
        shutil.move(src, dest)
    except OSError:
        pass

# ==========================================
# PHYSICAL SCHEMA SYNCHRONIZATION
# ==========================================
def safe_convert(val, app_type):
    """Safely cast user values to target database data types, strictly failing if it's destructive."""
    if val is None or str(val).strip() == "": return True, None
    try:
        if app_type == "boolean":
            v = int(float(val))
            if v in (0, 1): return True, v
            return False, None
        elif app_type == "int":
            return True, int(float(val))
        elif app_type == "float":
            return True, float(val)
        elif app_type in ("list", "matrix"):
            parsed = ast.literal_eval(str(val))
            if isinstance(parsed, list):
                return True, str(parsed)
            return False, None
        else:
            return True, str(val)
    except (ValueError, SyntaxError, TypeError):
        return False, None

def sync_physical_table(db_path, class_id, class_name, parent_widget=None):
    safe_table_name = f"objects_{sanitize_name(class_name)}"
    
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        
        cur.execute("SELECT name, data_type, show_in_table, is_title, lookup_query, is_unique, is_required FROM attributes WHERE class_id = ? ORDER BY row_order", (class_id,))
        attributes = cur.fetchall()

        required_cols = []
        attr_app_types = {}
        unique_flags = {}    # safe_col_name -> bool (enforced via UNIQUE index)
        required_flags = {}  # safe_col_name -> bool (enforced via NOT NULL when feasible)

        for attr_name, attr_type, show_in_table, is_title, lookup_query, is_unique, is_required in attributes:
            safe_col_name = sanitize_name(attr_name)
            if attr_type == "look-through":
                attr_app_types[safe_col_name] = attr_type
                if lookup_query:
                    parts = lookup_query.split('.')
                    if len(parts) == 2:
                        tgt_class = sanitize_name(parts[0].strip())
                        cur.execute(f"CREATE TABLE IF NOT EXISTS {qid('objects_' + tgt_class)} (id INTEGER PRIMARY KEY AUTOINCREMENT)")
                        cur.execute(f"CREATE VIEW IF NOT EXISTS {qid('base_view_objects_' + tgt_class)} AS SELECT id AS [ID] FROM {qid('objects_' + tgt_class)}")
                continue

            if attr_type in ("int", "boolean"): sql_type = "INTEGER"
            elif attr_type == "float": sql_type = "REAL"
            else: sql_type = "TEXT"

            required_cols.append((safe_col_name, sql_type))
            attr_app_types[safe_col_name] = attr_type
            unique_flags[safe_col_name] = bool(is_unique)
            required_flags[safe_col_name] = bool(is_required)

        def make_cols_def(notnull_map):
            return ", ".join([f"{qid(c)} {t}" + (" NOT NULL" if notnull_map.get(c) else "") for c, t in required_cols])

        cur.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name=?", (safe_table_name,))
        if cur.fetchone()[0] == 0:
            # Fresh, empty table: every required column can safely carry NOT NULL.
            desired_notnull = dict(required_flags)
            cols_def = make_cols_def(desired_notnull)
            cur.execute(f"CREATE TABLE {qid(safe_table_name)} (id INTEGER PRIMARY KEY AUTOINCREMENT{', ' + cols_def if cols_def else ''})")
        else:
            cur.execute(f"PRAGMA table_info({qid(safe_table_name)})")
            info_rows = cur.fetchall()
            existing_cols = {row[1]: row[2] for row in info_rows}
            existing_notnull = {row[1]: row[3] for row in info_rows}

            cur.execute(f"SELECT COUNT(*) FROM {qid(safe_table_name)}")
            table_has_rows = cur.fetchone()[0] > 0

            type_mismatches = []
            for col_name, sql_type in required_cols:
                if col_name in existing_cols and existing_cols[col_name] != sql_type:
                    type_mismatches.append(col_name)

            cols_to_remove = set(existing_cols.keys()) - {c for c, t in required_cols} - {'id'}

            # Decide the NOT NULL constraint we actually want per column, degrading to
            # nullable wherever enforcing it could break existing data (so a rebuild can
            # never fail on a NOT NULL violation).
            desired_notnull = {}
            notnull_changes = []
            for col_name, sql_type in required_cols:
                want = required_flags.get(col_name, False)
                if want:
                    if col_name not in existing_cols:
                        # New column: only enforceable while the table is still empty.
                        if table_has_rows: want = False
                    elif col_name in type_mismatches:
                        # Conversion may null out incompatible values.
                        want = False
                    elif table_has_rows:
                        cur.execute(f"SELECT 1 FROM {qid(safe_table_name)} WHERE {qid(col_name)} IS NULL LIMIT 1")
                        if cur.fetchone(): want = False
                desired_notnull[col_name] = want

                if col_name in existing_cols:
                    if bool(existing_notnull.get(col_name, 0)) != want:
                        notnull_changes.append(col_name)
                elif want:
                    # New required column on an empty table -> rebuild to create it NOT NULL
                    # (ALTER TABLE ADD COLUMN cannot add a NOT NULL column without a default).
                    notnull_changes.append(col_name)

            if type_mismatches or cols_to_remove or notnull_changes:
                conversion_failures = 0
                if type_mismatches:
                    type_mismatches_escaped = ", ".join([qid(c) for c in type_mismatches])
                    cur.execute(f"SELECT id, {type_mismatches_escaped} FROM {qid(safe_table_name)}")
                    rows = cur.fetchall()
                    
                    for row in rows:
                        for idx, col in enumerate(type_mismatches):
                            val = row[idx+1]
                            success, _ = safe_convert(val, attr_app_types[col])
                            if not success:
                                conversion_failures += 1
                                
                if conversion_failures > 0:
                    if parent_widget:
                        reply = QMessageBox.question(
                            parent_widget,
                            "Data Type Conversion Warning",
                            f"{conversion_failures} object values cannot be converted safely to the newly selected data types.\n"
                            f"If you continue, these incompatible values will be cleared (set to None).\n\n"
                            "Do you want to continue?",
                            QMessageBox.Yes | QMessageBox.No
                        )
                        if reply == QMessageBox.No: return None, None
                    else:
                        raise RuntimeError(f"Sync aborted: Data type conversion would cause {conversion_failures} object values to be cleared to None. Manual resolution required.")
                            
                cur.execute("PRAGMA foreign_keys")
                fk_state = cur.fetchone()[0]
                
                cur.execute("PRAGMA legacy_alter_table")
                legacy_row = cur.fetchone()
                legacy_state = legacy_row[0] if legacy_row else 0
                
                conn.commit()
                conn.execute("PRAGMA foreign_keys = OFF")
                
                try:
                    conn.execute("BEGIN IMMEDIATE")
                    
                    # Prevent modern SQLite from verifying/modifying views dynamically during ALTER TABLE 
                    conn.execute("PRAGMA legacy_alter_table = ON")
                    
                    # Clear views mapped to the old schema explicitly before renaming/dropping the underlying table
                    cur.execute(f"DROP VIEW IF EXISTS {qid('view_' + safe_table_name)}")
                    cur.execute(f"DROP VIEW IF EXISTS {qid('full_view_' + safe_table_name)}")
                    cur.execute(f"DROP VIEW IF EXISTS {qid('base_view_' + safe_table_name)}")
                    
                    new_table = f"new_{safe_table_name}"
                    cur.execute(f"DROP TABLE IF EXISTS {qid(new_table)}")

                    cols_def = make_cols_def(desired_notnull)
                    cur.execute(f"CREATE TABLE {qid(new_table)} (id INTEGER PRIMARY KEY AUTOINCREMENT{', ' + cols_def if cols_def else ''})")
                    
                    common_cols = list(set(existing_cols.keys()).intersection([c for c, t in required_cols]))
                    if common_cols:
                        escaped_common_cols = ", ".join([qid(c) for c in common_cols])
                        cur.execute(f"SELECT id, {escaped_common_cols} FROM {qid(safe_table_name)}")
                        old_data = cur.fetchall()
                        
                        for row in old_data:
                            row_id = row[0]
                            new_values = [row_id]
                            insert_cols = ["id"]
                            placeholders = ["?"]
                            
                            for idx, col in enumerate(common_cols):
                                val = row[idx+1]
                                new_val = val
                                if col in type_mismatches:
                                    success, conv_val = safe_convert(val, attr_app_types[col])
                                    new_val = conv_val if success else None
                                
                                new_values.append(new_val)
                                insert_cols.append(qid(col))
                                placeholders.append("?")
                                
                            cur.execute(f"INSERT INTO {qid(new_table)} ({', '.join(insert_cols)}) VALUES ({', '.join(placeholders)})", new_values)
                            
                    cur.execute(f"DROP TABLE {qid(safe_table_name)}")
                    cur.execute(f"ALTER TABLE {qid(new_table)} RENAME TO {qid(safe_table_name)}")
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    raise e
                finally:
                    conn.execute(f"PRAGMA legacy_alter_table = {legacy_state}")
                    conn.execute(f"PRAGMA foreign_keys = {fk_state}")
                
            else:
                for col_name, col_type in required_cols:
                    if col_name not in existing_cols:
                        cur.execute(f"ALTER TABLE {qid(safe_table_name)} ADD COLUMN {qid(col_name)} {col_type}")

        # --- Schema-level UNIQUE enforcement via standalone indexes ---
        # (ALTER TABLE cannot add a UNIQUE column, so unique constraints live in indexes
        # named ux_<table>_<col>. SQLite unique indexes treat multiple NULLs as distinct,
        # which matches the app's "uniqueness ignores empty values" behaviour.)
        ux_prefix = f"ux_{safe_table_name}_"
        cur.execute(f"PRAGMA index_list({qid(safe_table_name)})")
        existing_index_names = [r[1] for r in cur.fetchall()]
        desired_unique_cols = [c for c, _t in required_cols if unique_flags.get(c)]

        for idx_name in existing_index_names:
            if idx_name.startswith(ux_prefix) and idx_name[len(ux_prefix):] not in desired_unique_cols:
                cur.execute(f"DROP INDEX IF EXISTS {qid(idx_name)}")

        for c in desired_unique_cols:
            idx_name = ux_prefix + c
            if idx_name not in existing_index_names:
                try:
                    cur.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS {qid(idx_name)} ON {qid(safe_table_name)} ({qid(c)})")
                except (sqlite3.IntegrityError, sqlite3.OperationalError):
                    # Existing duplicate values (app-level checks normally prevent this).
                    # Skip schema-level enforcement rather than abort the whole sync.
                    pass

        # Outgoing relationships uses show_in_base
        cur.execute("SELECT c.name, r.rel_type, c.id, r.show_in_base FROM relationships r JOIN classes c ON r.target_class = c.id WHERE r.source_class = ? ORDER BY r.row_order", (class_id,))
        outgoing_rels = cur.fetchall()
        
        for target_name, rel_type, target_class_id, show_in_base in outgoing_rels:
            safe_target_name = sanitize_name(target_name)
            cur.execute(f"CREATE TABLE IF NOT EXISTS {qid('objects_' + safe_target_name)} (id INTEGER PRIMARY KEY AUTOINCREMENT)")
            cur.execute(f"CREATE VIEW IF NOT EXISTS {qid('base_view_objects_' + safe_target_name)} AS SELECT id AS [ID] FROM {qid('objects_' + safe_target_name)}")
            
            junc_table = f"rel_{safe_table_name}_to_objects_{safe_target_name}"
            cur.execute(f"CREATE TABLE IF NOT EXISTS {qid(junc_table)} (id INTEGER PRIMARY KEY AUTOINCREMENT, source_id INTEGER, target_id INTEGER, FOREIGN KEY(source_id) REFERENCES {qid(safe_table_name)}(id) ON DELETE CASCADE, FOREIGN KEY(target_id) REFERENCES {qid('objects_' + safe_target_name)}(id) ON DELETE CASCADE)")

        # Incoming relationships uses show_in_target
        cur.execute("SELECT c.name, r.rel_type, c.id, r.show_in_target FROM relationships r JOIN classes c ON r.source_class = c.id WHERE r.target_class = ? ORDER BY r.row_order", (class_id,))
        incoming_rels = cur.fetchall()

        for source_name, rel_type, source_class_id, show_in_target in incoming_rels:
            safe_source_name = sanitize_name(source_name)
            cur.execute(f"CREATE TABLE IF NOT EXISTS {qid('objects_' + safe_source_name)} (id INTEGER PRIMARY KEY AUTOINCREMENT)")
            cur.execute(f"CREATE VIEW IF NOT EXISTS {qid('base_view_objects_' + safe_source_name)} AS SELECT id AS [ID] FROM {qid('objects_' + safe_source_name)}")
            
            junc_table = f"rel_objects_{safe_source_name}_to_{safe_table_name}"
            cur.execute(f"CREATE TABLE IF NOT EXISTS {qid(junc_table)} (id INTEGER PRIMARY KEY AUTOINCREMENT, source_id INTEGER, target_id INTEGER, FOREIGN KEY(source_id) REFERENCES {qid('objects_' + safe_source_name)}(id) ON DELETE CASCADE, FOREIGN KEY(target_id) REFERENCES {qid(safe_table_name)}(id) ON DELETE CASCADE)")

        base_selects = ["m.id AS [ID]"]
        for attr_name, attr_type, show_in_table, is_title, lookup_query, is_unique, is_required in attributes:
            safe_col_name = sanitize_name(attr_name)
            if attr_type == "look-through":
                if lookup_query:
                    parts = lookup_query.split('.')
                    if len(parts) == 2:
                        tgt_class, tgt_attr = parts
                        safe_tgt_class = sanitize_name(tgt_class.strip())
                        junc_table = f"rel_{safe_table_name}_to_objects_{safe_tgt_class}"
                        # A look-through resolves through the relationship's junction table.
                        # If no such relationship exists (junction missing), fall back to NULL
                        # so existing data still loads instead of breaking the whole view.
                        cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (junc_table,))
                        if cur.fetchone():
                            subquery = f"(SELECT GROUP_CONCAT(tgt_v.{qid(tgt_attr.strip())}) FROM {qid(junc_table)} j JOIN {qid('base_view_objects_' + safe_tgt_class)} tgt_v ON j.target_id = tgt_v.[ID] WHERE j.source_id = m.id)"
                            base_selects.append(f"{subquery} AS {qid(attr_name)}")
                        else:
                            base_selects.append(f"NULL AS {qid(attr_name)}")
                    else:
                        raise ValueError(f"Invalid lookup format '{lookup_query}'. Expected 'TargetClass.Attribute'.")
                else:
                    base_selects.append(f"NULL AS {qid(attr_name)}")
            elif attr_type == "file":
                # The raw column holds "<uuid>__name"; expose only the display name so the
                # table, look-throughs and relationship titles show the clean filename.
                base_selects.append(f"substr(m.{qid(safe_col_name)}, {FILE_PREFIX_LEN + 1}) AS {qid(attr_name)}")
            else:
                base_selects.append(f"m.{qid(safe_col_name)} AS {qid(attr_name)}")

        base_view_name = f"base_view_{safe_table_name}"
        cur.execute(f"DROP VIEW IF EXISTS {qid(base_view_name)}")
        cur.execute(f"CREATE VIEW {qid(base_view_name)} AS SELECT {', '.join(base_selects)} FROM {qid(safe_table_name)} m")

        ui_selects = ["v.[ID]"]
        full_selects = ["v.[ID]"]

        for attr_name, attr_type, show_in_table, is_title, lookup_query, is_unique, is_required in attributes:
            col_sql = f"v.{qid(attr_name)}"
            if show_in_table: ui_selects.append(col_sql)
            full_selects.append(col_sql)

        for target_name, rel_type, target_class_id, show_in_base in outgoing_rels:
            safe_target_name = sanitize_name(target_name)
            junc_table = f"rel_{safe_table_name}_to_objects_{safe_target_name}"
            
            # ALWAYS process explicitly requested ID columns
            id_display_label = qid(f"{target_name} (IDs)")
            id_query_str = f"(SELECT GROUP_CONCAT(j.target_id) FROM {qid(junc_table)} j WHERE j.source_id = v.[ID]) AS {id_display_label}"
            if show_in_base: ui_selects.append(id_query_str)
            full_selects.append(id_query_str)
            
            # Process Title Columns in addition to the IDs
            cur.execute("SELECT name FROM attributes WHERE class_id = ? AND is_title = 1 ORDER BY row_order", (target_class_id,))
            title_rows = cur.fetchall()
            for (t_name,) in title_rows:
                display_label = qid(f"{target_name} ({t_name})")
                query_str = f"(SELECT GROUP_CONCAT(tgt_v.{qid(t_name)}) FROM {qid(junc_table)} j JOIN {qid('base_view_objects_' + safe_target_name)} tgt_v ON j.target_id = tgt_v.[ID] WHERE j.source_id = v.[ID]) AS {display_label}"
                if show_in_base: ui_selects.append(query_str)
                full_selects.append(query_str)

        for source_name, rel_type, source_class_id, show_in_target in incoming_rels:
            safe_source_name = sanitize_name(source_name)
            junc_table = f"rel_objects_{safe_source_name}_to_{safe_table_name}"
            
            # ALWAYS process explicitly requested ID columns
            id_display_label = qid(f"From {source_name} (IDs)")
            id_query_str = f"(SELECT GROUP_CONCAT(j.source_id) FROM {qid(junc_table)} j WHERE j.target_id = v.[ID]) AS {id_display_label}"
            if show_in_target: ui_selects.append(id_query_str)
            full_selects.append(id_query_str)
            
            # Process Title Columns in addition to the IDs
            cur.execute("SELECT name FROM attributes WHERE class_id = ? AND is_title = 1 ORDER BY row_order", (source_class_id,))
            title_rows = cur.fetchall()
            for (t_name,) in title_rows:
                display_label = qid(f"From {source_name} ({t_name})")
                query_str = f"(SELECT GROUP_CONCAT(src_v.{qid(t_name)}) FROM {qid(junc_table)} j JOIN {qid('base_view_objects_' + safe_source_name)} src_v ON j.source_id = src_v.[ID] WHERE j.target_id = v.[ID]) AS {display_label}"
                if show_in_target: ui_selects.append(query_str)
                full_selects.append(query_str)

        view_name = f"view_{safe_table_name}"
        cur.execute(f"DROP VIEW IF EXISTS {qid(view_name)}")
        cur.execute(f"CREATE VIEW {qid(view_name)} AS SELECT {', '.join(ui_selects)} FROM {qid(base_view_name)} v")

        full_view_name = f"full_view_{safe_table_name}"
        cur.execute(f"DROP VIEW IF EXISTS {qid(full_view_name)}")
        cur.execute(f"CREATE VIEW {qid(full_view_name)} AS SELECT {', '.join(full_selects)} FROM {qid(base_view_name)} v")
    
    return view_name, safe_table_name


class NullableDateTimeEdit(QWidget):
    """A date/time editor that can represent 'no value' (NULL).

    A checkbox toggles whether a date is set. New objects default to checked
    with the current date/time; unchecking stores NULL.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)

        self.check = QCheckBox()
        self.check.setToolTip("Tick to set a date/time, untick for no date (empty / NULL).")
        self.edit = QDateTimeEdit()
        self.edit.setCalendarPopup(True)
        self.edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.edit.setDateTime(QDateTime.currentDateTime())

        lay.addWidget(self.check)
        lay.addWidget(self.edit, 1)

        self.check.toggled.connect(self.edit.setEnabled)
        self.check.setChecked(True)
        self.edit.setEnabled(True)

    def value(self):
        """Return the formatted string, or None when no date is set."""
        if not self.check.isChecked():
            return None
        return self.edit.dateTime().toString("yyyy-MM-dd HH:mm:ss")

    def set_value(self, val):
        if val is None or str(val).strip() == "":
            self.check.setChecked(False)
            self.edit.setEnabled(False)
        else:
            dt = QDateTime.fromString(str(val), "yyyy-MM-dd HH:mm:ss")
            if dt.isValid():
                self.edit.setDateTime(dt)
            self.check.setChecked(True)
            self.edit.setEnabled(True)


class FileAttributeWidget(QWidget):
    """Attach / open / clear a single file for a 'file' attribute.

    Keeps the existing stored reference until the user picks a new file or clears it.
    The actual copy into storage happens at save time (transaction-safe), via plan().
    """
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self.current = None      # existing stored reference ("<uuid>__name") or None
        self.pending = None      # absolute path of a newly chosen file, or None
        self.cleared = False     # user explicitly cleared an existing file

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)

        self.label = QLabel()
        self.label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        self.btn_choose = QPushButton(" Choose...")
        self.btn_choose.setIcon(qta.icon('fa5s.folder-open'))
        self.btn_choose.clicked.connect(self._choose)

        self.btn_open = QPushButton()
        self.btn_open.setIcon(qta.icon('fa5s.external-link-alt'))
        self.btn_open.setToolTip("Open the current file")
        self.btn_open.setFixedWidth(32)
        self.btn_open.clicked.connect(self._open)

        self.btn_clear = QPushButton()
        self.btn_clear.setIcon(qta.icon('fa5s.times', color='#ff4c4c'))
        self.btn_clear.setToolTip("Remove the attached file")
        self.btn_clear.setFixedWidth(32)
        self.btn_clear.clicked.connect(self._clear)

        lay.addWidget(self.label, 1)
        lay.addWidget(self.btn_choose)
        lay.addWidget(self.btn_open)
        lay.addWidget(self.btn_clear)
        self._refresh()

    def set_existing(self, stored_value):
        self.current = stored_value or None
        self.pending = None
        self.cleared = False
        self._refresh()

    def _refresh(self):
        if self.pending:
            self.label.setText(f"<b>{os.path.basename(self.pending)}</b> <span style='color:#888;'>(new)</span>")
            self.btn_open.setEnabled(True)
            self.btn_clear.setEnabled(True)
            return
        if self.current and not self.cleared:
            name = display_file_name(self.current)
            path = resolve_file_path(self.db_path, self.current)
            if path and os.path.exists(path):
                self.label.setText(name)
            else:
                self.label.setText(f"{name} <span style='color:#cc8800;'>&#9888; missing</span>")
                self.label.setToolTip("The stored file is no longer in the files folder.")
            self.btn_open.setEnabled(True)
            self.btn_clear.setEnabled(True)
            return
        self.label.setText("<span style='color:#888;'>No file</span>")
        self.btn_open.setEnabled(False)
        self.btn_clear.setEnabled(False)

    def _choose(self):
        path, _ = QFileDialog.getOpenFileName(self, "Choose a file to attach", "", "All Files (*.*)")
        if path:
            self.pending = path
            self.cleared = False
            self._refresh()

    def _clear(self):
        self.pending = None
        self.cleared = True
        self._refresh()

    def _open(self):
        target = self.pending or (resolve_file_path(self.db_path, self.current) if (self.current and not self.cleared) else None)
        if target and os.path.exists(target):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(target)))
        else:
            QMessageBox.warning(self, "File Not Found", "The file could not be found on disk.")

    def is_empty(self):
        if self.pending:
            return False
        return self.cleared or not self.current

    def plan(self):
        """Return the save action: ('new', src, new_stored, old) | ('clear', old) | ('keep', current)."""
        if self.pending:
            new_stored = make_stored_filename(self.pending)
            return ('new', self.pending, new_stored, self.current)
        if self.cleared:
            return ('clear', self.current)
        return ('keep', self.current)


class ObjectEditorDialog(QDialog):
    def __init__(self, db_path, class_id, class_name, table_name, obj_id=None, parent=None):
        super().__init__(parent)
        self.obj_id = obj_id
        mode_text = "Edit" if obj_id else "Add New"
        self.setWindowTitle(f"{mode_text} {class_name}")
        self.setWindowIcon(get_app_icon())
        self.resize(500, 550) 
        
        self.db_path = db_path
        self.class_id = class_id
        self.table_name = table_name
        self.input_widgets = {} 
        self.rel_widgets = {}   
        self.attr_app_types = {} 
        self.attr_constraints = {}
        self.matrix_col_counts = {}
        
        layout = QVBoxLayout(self)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        
        scroll_widget = QWidget()
        form_layout = QFormLayout(scroll_widget)
        form_layout.setContentsMargins(10, 10, 10, 10)
        form_layout.setSpacing(15)
        
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            
            cur.execute("SELECT id, name, data_type, is_unique, is_required, lookup_query FROM attributes WHERE class_id = ? ORDER BY row_order", (self.class_id,))
            attributes = cur.fetchall()

            for attr_id, attr_name, attr_type, is_unique, is_required, lookup_query in attributes:
                safe_col_name = sanitize_name(attr_name)
                
                if attr_type == "look-through": continue
                    
                self.attr_app_types[safe_col_name] = attr_type
                self.attr_constraints[safe_col_name] = {
                    'name': attr_name,
                    'unique': bool(is_unique),
                    'required': bool(is_required)
                }
                
                if attr_type == "int":
                    # Plain text + validator instead of QSpinBox: QSpinBox is limited to
                    # 32-bit and silently clamps, whereas SQLite INTEGER is 64-bit. Parsing
                    # in Python keeps the full range with no silent data loss.
                    widget = QLineEdit()
                    widget.setValidator(QRegularExpressionValidator(QRegularExpression(r'^-?\d*$')))
                    widget.setPlaceholderText("Whole number")
                elif attr_type == "float":
                    # Likewise avoids QDoubleSpinBox's range cap and 4-decimal rounding.
                    widget = QLineEdit()
                    widget.setValidator(QRegularExpressionValidator(QRegularExpression(r'^-?\d*\.?\d*([eE][-+]?\d+)?$')))
                    widget.setPlaceholderText("Number")
                elif attr_type == "boolean":
                    widget = QCheckBox("Yes / True")
                elif attr_type == "date":
                    widget = NullableDateTimeEdit()
                elif attr_type == "file":
                    widget = FileAttributeWidget(self.db_path)
                elif attr_type == "discrete":
                    widget = QComboBox()
                    widget.addItem("(none)", None)  # blank choice; required check rejects it
                    try:
                        type_id = int(lookup_query)
                        for (opt,) in cur.execute("SELECT value FROM discrete_options WHERE type_id = ? ORDER BY row_order", (type_id,)).fetchall():
                            widget.addItem(opt, opt)
                    except (ValueError, TypeError, sqlite3.OperationalError):
                        pass
                elif attr_type == "long string":
                    widget = QTextEdit()
                    widget.setMaximumHeight(100) 
                elif attr_type in ("list", "matrix"):
                    widget = QLineEdit("[]") 
                    if attr_type == "matrix":
                        cur.execute("SELECT count(*) FROM matrix_columns WHERE attribute_id = ?", (attr_id,))
                        col_count = cur.fetchone()[0]
                        self.matrix_col_counts[safe_col_name] = col_count
                        widget.setToolTip(f"Provide exactly {col_count} list(s) inside the main list. e.g. [[1,2], [3,4]]")
                    else:
                        widget.setToolTip("Must be a valid python list e.g. ['A', 'B'] or [1, 2]")
                else:
                    widget = QLineEdit() 
                    
                self.input_widgets[safe_col_name] = widget
                label_text = attr_name + (" <span style='color:red;'>*</span>" if is_required else "") + ":"
                form_layout.addRow(QLabel(label_text), widget)

            cur.execute("""
                SELECT c.id, c.name, r.rel_type 
                FROM relationships r JOIN classes c ON r.target_class = c.id 
                WHERE r.source_class = ? ORDER BY r.row_order
            """, (self.class_id,))
            relationships = cur.fetchall()

            for target_class_id, target_name, rel_type in relationships:
                safe_target_name = sanitize_name(target_name)
                junc_table = f"rel_{self.table_name}_to_objects_{safe_target_name}"
                
                rel_container = QVBoxLayout()
                rel_container.setContentsMargins(0, 0, 0, 0)
                rel_container.setSpacing(5)
                
                search_box = QLineEdit()
                search_box.setPlaceholderText(f"Search {target_name}...")
                rel_container.addWidget(search_box)
                
                list_widget = QListWidget()
                if rel_type == "one_to_many":
                    list_widget.setSelectionMode(QListWidget.SingleSelection)
                else:
                    list_widget.setSelectionMode(QListWidget.MultiSelection)
                list_widget.setMinimumHeight(120) 
                
                cur.execute("SELECT name FROM attributes WHERE class_id = ? AND is_title = 1 ORDER BY row_order", (target_class_id,))
                title_rows = cur.fetchall()
                
                try:
                    if title_rows:
                        title_cols = [r[0] for r in title_rows] 
                        escaped_cols = ", ".join([qid(c) for c in title_cols])
                        cur.execute(f"SELECT [ID], {escaped_cols} FROM {qid('base_view_objects_' + safe_target_name)}")
                        
                        for row in cur.fetchall():
                            t_id = row[0]
                            display_vals = ["None" if val is None or str(val).strip() == "" else str(val) for val in row[1:]]
                            combined_titles = " --- ".join(display_vals)
                            display_text = f"{combined_titles} (ID: {t_id})"
                                
                            item = QListWidgetItem(display_text)
                            item.setData(Qt.UserRole, t_id)
                            item.setData(Qt.UserRole + 1, combined_titles.lower())
                            list_widget.addItem(item)
                    else:
                        cur.execute(f"SELECT [ID] FROM {qid('base_view_objects_' + safe_target_name)}")
                        for row in cur.fetchall():
                            t_id = row[0]
                            item = QListWidgetItem(f"{target_name} #{t_id}")
                            item.setData(Qt.UserRole, t_id)
                            item.setData(Qt.UserRole + 1, target_name.lower()) 
                            list_widget.addItem(item)
                except sqlite3.OperationalError:
                    pass 
                    
                rel_container.addWidget(list_widget)
                search_box.textChanged.connect(lambda text, lw=list_widget: self.filter_list_items(text, lw))
                    
                self.rel_widgets[junc_table] = {"widget": list_widget, "rel_type": rel_type, "target_name": target_name}
                form_layout.addRow(f"Rel: {target_name}:", rel_container)

            if self.obj_id:
                cols = list(self.input_widgets.keys())
                if cols:
                    escaped_cols = [qid(c) for c in cols]
                    cur.execute(f"SELECT {', '.join(escaped_cols)} FROM {qid(self.table_name)} WHERE id = ?", (self.obj_id,))
                    row = cur.fetchone()
                    if row:
                        for idx, col in enumerate(cols):
                            val = row[idx]
                            widget = self.input_widgets[col]

                            if isinstance(widget, NullableDateTimeEdit):
                                # Always call (even for None) so a stored NULL unticks the box.
                                widget.set_value(val)
                            elif isinstance(widget, FileAttributeWidget):
                                widget.set_existing(val)
                            elif isinstance(widget, QComboBox):  # discrete
                                pos = widget.findData(val) if val is not None else 0
                                if pos < 0:
                                    # Value no longer in the option set: show it so we don't lose it.
                                    widget.addItem(str(val), val)
                                    pos = widget.count() - 1
                                widget.setCurrentIndex(pos)
                            elif val is not None:
                                if isinstance(widget, QCheckBox):
                                    widget.setChecked(bool(val))
                                elif isinstance(widget, QTextEdit):
                                    widget.setPlainText(str(val))
                                else:
                                    widget.setText(str(val))
                                    
                for junc_table, data in self.rel_widgets.items():
                    cur.execute(f"SELECT target_id FROM {qid(junc_table)} WHERE source_id = ?", (self.obj_id,))
                    existing_targets = [r[0] for r in cur.fetchall()]
                    lw = data["widget"]
                    for i in range(lw.count()):
                        item = lw.item(i)
                        if item.data(Qt.UserRole) in existing_targets: item.setSelected(True)

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        btn_layout = QHBoxLayout()
        btn_save = QPushButton(" Save Object")
        btn_save.setIcon(qta.icon('fa5s.save'))
        btn_save.clicked.connect(self.save_record)
        
        btn_cancel = QPushButton(" Cancel")
        btn_cancel.setIcon(qta.icon('fa5s.times', color='#ff4c4c'))
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addStretch() 
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

    def filter_list_items(self, search_text, list_widget):
        search_text = search_text.lower()
        for i in range(list_widget.count()):
            item = list_widget.item(i)
            search_target = item.data(Qt.UserRole + 1)
            if search_target is None: search_target = item.text().lower() 
            item.setHidden(search_text not in search_target)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter: return 
        super().keyPressEvent(event)

    def save_record(self):
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("PRAGMA foreign_keys = 1")
                cur = conn.cursor()
                
                columns = []
                values = []
                placeholders = []
                file_copies = []      # (source_path, stored_name) to copy into storage on commit
                files_to_trash = []   # old stored names to trash after a successful commit

                for col_name, widget in self.input_widgets.items():
                    app_type = self.attr_app_types.get(col_name)
                    constraints = self.attr_constraints.get(col_name)
                    attr_display_name = constraints['name']

                    if isinstance(widget, FileAttributeWidget):
                        action = widget.plan()
                        is_empty = widget.is_empty()
                        if action[0] == 'new':
                            _, src, new_stored, old = action
                            val = new_stored
                            file_copies.append((src, new_stored))
                            if old:
                                files_to_trash.append(old)
                        elif action[0] == 'clear':
                            val = None
                            if action[1]:
                                files_to_trash.append(action[1])
                        else:  # keep
                            val = action[1]
                    elif isinstance(widget, QComboBox):  # discrete attribute
                        val = widget.currentData()
                        is_empty = (val is None)
                    elif isinstance(widget, QCheckBox):
                        val = 1 if widget.isChecked() else 0
                        is_empty = False
                    elif isinstance(widget, NullableDateTimeEdit):
                        val = widget.value()
                        is_empty = (val is None)
                    elif isinstance(widget, QTextEdit):
                        raw_text = widget.toPlainText().strip()
                        is_empty = (raw_text == "")
                        val = None if is_empty else raw_text
                    else:
                        raw_text = widget.text().strip()
                        is_empty = (raw_text == "")
                        if is_empty:
                            val = None
                        else:
                            if app_type in ("list", "matrix"):
                                try:
                                    parsed = ast.literal_eval(raw_text)
                                    if not isinstance(parsed, list): raise ValueError("Not a list.")

                                    if app_type == "matrix":
                                        expected_cols = self.matrix_col_counts.get(col_name, 0)
                                        if len(parsed) != expected_cols:
                                            raise ValueError(f"Matrix expects exactly {expected_cols} column list(s) inside the main list.")
                                        for inner in parsed:
                                            if not isinstance(inner, list): raise ValueError("Each matrix column must be a list.")
                                    val = str(parsed)
                                except Exception as e:
                                    err_msg = str(e) if str(e) else "Invalid Python syntax."
                                    QMessageBox.warning(self, "Validation Error", f"'{attr_display_name}' parsing failed: {err_msg}\n\nExample: [['A', 'B'], [1, 2]]")
                                    return
                            elif app_type == "int":
                                try:
                                    val = int(raw_text)
                                except ValueError:
                                    QMessageBox.warning(self, "Validation Error", f"'{attr_display_name}' must be a whole number.")
                                    return
                            elif app_type == "float":
                                try:
                                    val = float(raw_text)
                                except ValueError:
                                    QMessageBox.warning(self, "Validation Error", f"'{attr_display_name}' must be a number.")
                                    return
                            else: val = raw_text

                    if constraints['required'] and is_empty:
                        QMessageBox.warning(self, "Validation Error", f"'{attr_display_name}' is a required field.")
                        return

                    if constraints['unique'] and not is_empty:
                        if self.obj_id:
                            cur.execute(f"SELECT id FROM {qid(self.table_name)} WHERE {qid(col_name)} = ? AND id != ?", (val, self.obj_id))
                        else:
                            cur.execute(f"SELECT id FROM {qid(self.table_name)} WHERE {qid(col_name)} = ?", (val,))
                            
                        if cur.fetchone():
                            QMessageBox.warning(self, "Validation Error", f"'{attr_display_name}' must be unique. The value '{val}' already exists.")
                            return

                    columns.append(col_name)
                    placeholders.append("?")
                    values.append(val)

                copied_abs = []
                try:
                    # Copy new attachments into storage first; if the DB write fails we
                    # delete them again so nothing is stranded.
                    if file_copies:
                        files_dir = files_dir_for(self.db_path, create=True)
                        for src, dest_name in file_copies:
                            dest_abs = os.path.join(files_dir, dest_name)
                            shutil.copy2(src, dest_abs)
                            copied_abs.append(dest_abs)

                    conn.execute("BEGIN IMMEDIATE")
                    active_obj_id = self.obj_id

                    if columns:
                        if self.obj_id:
                            set_clause = ", ".join([f"{qid(c)} = ?" for c in columns])
                            query = f"UPDATE {qid(self.table_name)} SET {set_clause} WHERE id = ?"
                            cur.execute(query, tuple(values) + (self.obj_id,))
                        else:
                            escaped_columns = [qid(c) for c in columns]
                            query = f"INSERT INTO {qid(self.table_name)} ({', '.join(escaped_columns)}) VALUES ({', '.join(placeholders)})"
                            cur.execute(query, tuple(values))
                            active_obj_id = cur.lastrowid
                    else:
                        if not self.obj_id:
                            cur.execute(f"INSERT INTO {qid(self.table_name)} DEFAULT VALUES")
                            active_obj_id = cur.lastrowid
                        
                    for junc_table, data in self.rel_widgets.items():
                        if self.obj_id:
                            cur.execute(f"DELETE FROM {qid(junc_table)} WHERE source_id = ?", (active_obj_id,))
                        selected_ids = [int(item.data(Qt.UserRole)) for item in data["widget"].selectedItems()]
                        for target_id in selected_ids:
                            cur.execute(f"INSERT INTO {qid(junc_table)} (source_id, target_id) VALUES (?, ?)", (active_obj_id, target_id))
                    
                    conn.commit()
                    # Commit succeeded: retire replaced/cleared files to the trash folder.
                    for old in files_to_trash:
                        trash_stored_file(self.db_path, old)
                    self.accept()
                except Exception as e:
                    conn.rollback()
                    for path in copied_abs:
                        try:
                            os.remove(path)
                        except OSError:
                            pass
                    raise e

        except Exception as e:
            QMessageBox.critical(self, "Database Error", str(e))


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Database Settings")
        self.setWindowIcon(get_app_icon())
        self.resize(700, 250)
        self.settings = QSettings("MyCompany", "DatabaseManagerApp")
        
        main_layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        db_tab = QWidget()
        db_layout = QVBoxLayout(db_tab)
        db_layout.addWidget(QLabel("Database Path:"))

        path_layout = QHBoxLayout()
        self.path_input = QLineEdit()
        self.path_input.setText(self.settings.value("db_path", ""))
        
        btn_browse = QPushButton(" Browse...")
        btn_browse.setIcon(qta.icon('fa5s.folder-open'))
        btn_browse.clicked.connect(self.browse_file)
        
        path_layout.addWidget(self.path_input)
        path_layout.addWidget(btn_browse)
        db_layout.addLayout(path_layout)
        
        btn_db_layout = QHBoxLayout()
        btn_db_layout.addStretch()
        btn_db_save = QPushButton(" Save")
        btn_db_save.setIcon(qta.icon('fa5s.save'))
        btn_db_save.clicked.connect(self.save_db_settings)
        btn_db_layout.addWidget(btn_db_save)
        
        db_layout.addStretch()
        db_layout.addLayout(btn_db_layout)
        self.tabs.addTab(db_tab, "Database")

        mod_tab = QWidget()
        mod_layout = QVBoxLayout(mod_tab)
        mod_layout.addWidget(QLabel("Module Output File Path:"))
        
        out_layout = QHBoxLayout()
        self.output_input = QLineEdit()
        
        db_path = self.settings.value("db_path", "")
        if db_path and os.path.exists(os.path.dirname(db_path)):
            default_out = os.path.join(os.path.dirname(db_path), "module_output.txt")
        else:
            default_out = os.path.join(os.path.expanduser("~"), "module_output.txt")
            
        self.output_input.setText(self.settings.value("module_output_path", default_out))
        
        btn_browse_out = QPushButton(" Browse...")
        btn_browse_out.setIcon(qta.icon('fa5s.folder-open'))
        btn_browse_out.clicked.connect(self.browse_output_file)
        
        out_layout.addWidget(self.output_input)
        out_layout.addWidget(btn_browse_out)
        mod_layout.addLayout(out_layout)
        
        btn_mod_layout = QHBoxLayout()
        btn_mod_layout.addStretch()
        btn_mod_save = QPushButton(" Save")
        btn_mod_save.setIcon(qta.icon('fa5s.save'))
        btn_mod_save.clicked.connect(self.save_module_path)
        btn_mod_layout.addWidget(btn_mod_save)
        
        mod_layout.addStretch()
        mod_layout.addLayout(btn_mod_layout)
        self.tabs.addTab(mod_tab, "Modules")

    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Select SQLite Database", "", "SQLite DB (*.db *.sqlite)")
        if file_name: self.path_input.setText(file_name)

    def browse_output_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Select Output File", "", "Text Files (*.txt)")
        if file_name: self.output_input.setText(file_name)

    def save_db_settings(self):
        path = self.path_input.text().strip()
        if not path: return
        if not os.path.exists(path): init_db(path)
        self.settings.setValue("db_path", path)
        QMessageBox.information(self, "Success", "Database path saved as default.")
        
    def save_module_path(self):
        self.settings.setValue("module_output_path", self.output_input.text().strip())
        QMessageBox.information(self, "Success", "Module output path saved.")


class ModuleBuilderDialog(QDialog):
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Module Builder")
        self.setWindowIcon(get_app_icon())
        self.resize(1000, 700)
        self.db_path = db_path
        self.current_module_id = None
        
        if pd is None:
            QMessageBox.warning(self, "Missing Library", "The 'pandas' library is required to use the Module Builder and return DataFrames.\n\nPlease install it via terminal: pip install pandas")
        
        main_layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)
        
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        self.module_list_widget = QTreeWidget()
        self.module_list_widget.setHeaderHidden(True)
        self.module_list_widget.itemClicked.connect(self.load_module)
        
        btn_add = QPushButton(" Create New Module")
        btn_add.setIcon(qta.icon('fa5s.plus-circle'))
        btn_add.clicked.connect(self.create_new_module)
        
        left_layout.addWidget(QLabel("<b>Saved Modules</b>"))
        left_layout.addWidget(self.module_list_widget)
        left_layout.addWidget(btn_add)
        
        self.editor_widget = QWidget()
        self.editor_layout = QVBoxLayout(self.editor_widget)
        self.editor_widget.setEnabled(False)
        
        info_layout = QHBoxLayout()
        self.module_name_input = QLineEdit()
        self.module_name_input.setPlaceholderText("Module Name")
        self.module_path_input = QLineEdit()
        self.module_path_input.setPlaceholderText("e.g. Reports/Inventory")
        info_layout.addWidget(QLabel("Name:"))
        info_layout.addWidget(self.module_name_input)
        info_layout.addWidget(QLabel("Path:"))
        info_layout.addWidget(self.module_path_input)
        self.editor_layout.addLayout(info_layout)
        
        self.editor = QPlainTextEdit()
        font = self.editor.font()
        font.setFamily("Courier New")
        font.setPointSize(11)
        font.setStyleHint(QFont.Monospace)
        self.editor.setFont(font)
        
        self.editor_layout.addWidget(QLabel("Python Script Editor:"))
        self.editor_layout.addWidget(self.editor)
        
        btn_layout = QHBoxLayout()
        
        btn_delete = QPushButton(" Delete Module")
        btn_delete.setIcon(qta.icon('fa5s.trash-alt', color='white'))
        btn_delete.setStyleSheet("background-color: #ff4c4c; color: white;")
        btn_delete.clicked.connect(self.delete_module)
        
        btn_save = QPushButton(" Save Module")
        btn_save.setIcon(qta.icon('fa5s.save'))
        btn_save.clicked.connect(self.save_module)
        
        btn_run = QPushButton(" Run Script")
        btn_run.setIcon(qta.icon('fa5s.play', color='white'))
        btn_run.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 5px 15px;")
        btn_run.clicked.connect(self.run_script)
        
        btn_layout.addWidget(btn_delete)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_run)
        self.editor_layout.addLayout(btn_layout)

        splitter.addWidget(left_widget)
        splitter.addWidget(self.editor_widget)
        splitter.setSizes([250, 750])
        
        self.refresh_module_list()

    def get_all_modules(self):
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT id, name, path FROM modules ORDER BY path ASC, name ASC")
            return cur.fetchall()

    def refresh_module_list(self):
        self.module_list_widget.clear()
        root_items = {}
        parent_item = self.module_list_widget.invisibleRootItem()
        
        for mid, mname, mpath in self.get_all_modules():
            mpath = (mpath or "").strip().strip('/')
            current_parent = parent_item
            
            if mpath:
                current_path = ""
                for part in mpath.split('/'):
                    part = part.strip()
                    if not part: continue
                    current_path = f"{current_path}/{part}" if current_path else part
                    
                    if current_path not in root_items:
                        folder_item = QTreeWidgetItem([part])
                        folder_item.setIcon(0, qta.icon('fa5s.folder', color='#FFC107'))
                        current_parent.addChild(folder_item)
                        root_items[current_path] = folder_item
                    
                    current_parent = root_items[current_path]
            
            item = QTreeWidgetItem([mname])
            item.setIcon(0, qta.icon('fa5s.file-code', color='#2196F3'))
            item.setData(0, Qt.UserRole, mid)
            current_parent.addChild(item)
            
        self.module_list_widget.expandAll()

    def create_new_module(self):
        self.current_module_id = None
        self.module_name_input.clear()
        self.module_path_input.clear()
        placeholder = (
            "# Write your Python script here.\n"
            "# Built-in Helper API:\n"
            "#   df = get_objects('ClassName') -> Returns a pandas DataFrame of the class view.\n\n"
            "print('Hello from Module Builder!')\n"
        )
        self.editor.setPlainText(placeholder)
        self.editor_widget.setEnabled(True)

    def load_module(self, item, column=0):
        mid = item.data(0, Qt.UserRole)
        if not mid: return
        
        self.current_module_id = mid
        self.editor_widget.setEnabled(True)
        
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT name, path, code FROM modules WHERE id = ?", (mid,))
            row = cur.fetchone()
            
            if row:
                self.module_name_input.setText(row[0])
                self.module_path_input.setText(row[1] if row[1] else "")
                self.editor.setPlainText(row[2] if row[2] else "")

    def save_module(self):
        name = self.module_name_input.text().strip()
        path_val = self.module_path_input.text().strip()
        code = self.editor.toPlainText()
        
        if not name:
            QMessageBox.warning(self, "Error", "Module name cannot be empty.")
            return
            
        try:
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                if self.current_module_id is None:
                    cur.execute("INSERT INTO modules (name, path, code) VALUES (?, ?, ?)", (name, path_val, code))
                    self.current_module_id = cur.lastrowid
                else:
                    cur.execute("UPDATE modules SET name = ?, path = ?, code = ? WHERE id = ?", (name, path_val, code, self.current_module_id))
                conn.commit()
            
            self.refresh_module_list()
            QMessageBox.information(self, "Success", "Module saved successfully!")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Error", "A module with this name already exists.")

    def delete_module(self):
        if self.current_module_id is None: return
        reply = QMessageBox.question(self, "Delete", "Are you sure you want to delete this module?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("DELETE FROM modules WHERE id = ?", (self.current_module_id,))
                conn.commit()
                
            self.current_module_id = None
            self.editor_widget.setEnabled(False)
            self.module_name_input.clear()
            self.module_path_input.clear()
            self.editor.clear()
            self.refresh_module_list()

    def get_objects(self, class_name):
        if pd is None: raise ImportError("Pandas library is not installed.")
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM classes WHERE name = ?", (class_name,))
            row = cur.fetchone()
            if not row:
                raise ValueError(f"Class '{class_name}' not found in the database.")
            cls_id = row[0]
            
        view_name, safe_table_name = sync_physical_table(self.db_path, cls_id, class_name, parent_widget=None)
        if not view_name: raise RuntimeError(f"Failed to sync schema for class '{class_name}'.")
            
        full_view_name = f"full_view_{safe_table_name}"
        with sqlite3.connect(self.db_path) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {qid(full_view_name)}", conn)
        return df

    def run_script(self):
        code = self.editor.toPlainText()
        settings = QSettings("MyCompany", "DatabaseManagerApp")
        out_path = settings.value("module_output_path", os.path.join(os.path.expanduser("~"), "module_output.txt"))
        context = {'get_objects': self.get_objects, 'pd': pd}
        
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(f"--- Script Executed at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n\n")
                with redirect_stdout(f), redirect_stderr(f): exec(code, context)
            QMessageBox.information(self, "Success", f"Script executed successfully.\nOutput saved to:\n{out_path}")
        except Exception as e:
            with open(out_path, 'a', encoding='utf-8') as f:
                f.write("\n\n--- RUNTIME ERROR ---\n")
                traceback.print_exc(file=f)
            QMessageBox.warning(self, "Script Error", f"An error occurred during execution.\nCheck the output file for details:\n{out_path}")


class ExportOptionsDialog(QDialog):
    """Lets the user pick what to export and in which shape."""
    def __init__(self, parent=None, has_filter=False):
        super().__init__(parent)
        self.setWindowTitle("Export Options")
        self.setWindowIcon(get_app_icon())
        self.resize(440, 360)

        layout = QVBoxLayout(self)

        layout_box = QGroupBox("Layout")
        lb = QVBoxLayout(layout_box)
        self.rb_layout_import = QRadioButton("Import-ready (round-trip)")
        self.rb_layout_import.setChecked(True)
        self.rb_layout_import.setToolTip("Attribute names as headers + relationship IDs. This file can be re-imported.")
        self.rb_layout_full = QRadioButton("Full display view (for reading)")
        self.rb_layout_full.setToolTip("Everything currently visible, including titles and incoming links. Not re-importable.")
        lb.addWidget(self.rb_layout_import)
        lb.addWidget(self.rb_layout_full)
        layout.addWidget(layout_box)

        rows_box = QGroupBox("Rows")
        rb = QVBoxLayout(rows_box)
        self.rb_rows_all = QRadioButton("All rows")
        self.rb_rows_all.setChecked(True)
        self.rb_rows_filter = QRadioButton("Current filter / search only")
        self.rb_rows_filter.setEnabled(has_filter)
        if not has_filter:
            self.rb_rows_filter.setToolTip("No active search filter.")
        rb.addWidget(self.rb_rows_all)
        rb.addWidget(self.rb_rows_filter)
        layout.addWidget(rows_box)

        fmt_box = QGroupBox("Format")
        fb = QHBoxLayout(fmt_box)
        self.rb_fmt_xlsx = QRadioButton("Excel (.xlsx)")
        self.rb_fmt_xlsx.setChecked(True)
        self.rb_fmt_csv = QRadioButton("CSV (.csv)")
        fb.addWidget(self.rb_fmt_xlsx)
        fb.addWidget(self.rb_fmt_csv)
        layout.addWidget(fmt_box)

        self.cb_template = QCheckBox("Template only (column headers, no data)")
        self.cb_template.setToolTip("Produce an empty file with the correct headers to fill in and import.")
        layout.addWidget(self.cb_template)

        buttons = QDialogButtonBox()
        btn_export = buttons.addButton(" Export...", QDialogButtonBox.AcceptRole)
        btn_export.setIcon(qta.icon('fa5s.file-export'))
        buttons.addButton(QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addStretch()
        layout.addWidget(buttons)

    def get_options(self):
        return {
            'import_ready': self.rb_layout_import.isChecked(),
            'rows_filtered': self.rb_rows_filter.isChecked() and self.rb_rows_filter.isEnabled(),
            'fmt': 'xlsx' if self.rb_fmt_xlsx.isChecked() else 'csv',
            'template': self.cb_template.isChecked(),
        }


class ImportPreviewDialog(QDialog):
    """Shows how the chosen file maps onto the class before anything is written."""
    def __init__(self, parent, analysis):
        super().__init__(parent)
        self.setWindowTitle("Import Preview")
        self.setWindowIcon(get_app_icon())
        self.resize(560, 480)

        layout = QVBoxLayout(self)
        report = QTextBrowser()
        report.setOpenExternalLinks(False)

        html = [f"<p><b>File:</b> {analysis['file']}<br><b>Data rows:</b> {analysis['data_count']}</p>"]

        if analysis['problems']:
            html.append("<p style='color:#c0392b;'><b>Cannot import yet:</b></p><ul>")
            for p in analysis['problems']:
                html.append(f"<li style='color:#c0392b;'>{p}</li>")
            html.append("</ul>")

        html.append("<p><b>Columns detected:</b></p><ul>")
        kind_color = {"ok": "#2e7d32", "key": "#1565c0", "ignore": "#888888"}
        for disp, label, kind in analysis['columns']:
            color = kind_color.get(kind, "#000000")
            html.append(f"<li><b>{disp}</b> &rarr; <span style='color:{color};'>{label}</span></li>")
        html.append("</ul>")

        html.append(
            "<p style='color:#555;'>Rows with a value in the <b>ID</b> column update the matching "
            "object; rows with an empty ID are added as new. Columns marked <i>ignored</i> are skipped. "
            "Relationship cells accept comma-separated IDs (or a single-title name).</p>"
        )
        report.setHtml("".join(html))
        layout.addWidget(report)

        buttons = QDialogButtonBox()
        self.btn_import = buttons.addButton(" Import", QDialogButtonBox.AcceptRole)
        self.btn_import.setIcon(qta.icon('fa5s.file-import'))
        self.btn_import.setEnabled(analysis['can_import'])
        buttons.addButton(QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)


class DataBrowserPage(QWidget):
    def __init__(self, main_window): 
        super().__init__()
        self.main_window = main_window
        self.current_class_id = None
        self.current_class_name = None
        
        self.current_view_name = None
        self.current_table_name = None
        self.db_path = ""
        self.hidden_columns_memory = {}
        self.file_columns = {}  # display header -> safe column name, for 'file' attributes
        
        self.page_size = 100
        self.current_page = 0
        self.total_pages = 0
        self.total_records = 0
        self.current_sort_col = "ID"
        self.current_sort_order = "ASC"
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(9, 9, 9, 9)

        self.header_label = QLabel("<h2>&nbsp;</h2>")
        layout.addWidget(self.header_label)

        search_layout = QHBoxLayout()
        self.col_combo = QComboBox()
        self.col_combo.addItem("All Columns", -1)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Filter text...")
        self.search_input.returnPressed.connect(self.trigger_search)

        # VS Code-style search toggles.
        self.btn_match_case = QPushButton("Aa")
        self.btn_match_case.setCheckable(True)
        self.btn_match_case.setFixedWidth(34)
        self.btn_match_case.setToolTip("Match Case")
        self.btn_match_case.toggled.connect(self.trigger_search)

        self.btn_whole_word = QPushButton("ab|")
        self.btn_whole_word.setCheckable(True)
        self.btn_whole_word.setFixedWidth(34)
        self.btn_whole_word.setToolTip("Match Whole Word")
        self.btn_whole_word.toggled.connect(self.trigger_search)

        self.btn_search = QPushButton(" Search")
        self.btn_search.setIcon(qta.icon('fa5s.search'))
        self.btn_search.clicked.connect(self.trigger_search)
        
        self.btn_import = QPushButton(" Import")
        self.btn_import.setIcon(qta.icon('fa5s.file-import'))
        self.btn_import.clicked.connect(self.import_data)
        
        self.btn_export = QPushButton(" Export")
        self.btn_export.setIcon(qta.icon('fa5s.file-export'))
        self.btn_export.clicked.connect(self.export_data)
        
        self.btn_add = QPushButton(" Add Object")
        self.btn_add.setIcon(qta.icon('fa5s.plus'))
        
        self.btn_edit = QPushButton(" Edit")
        self.btn_edit.setIcon(qta.icon('fa5s.edit'))
        
        self.btn_delete = QPushButton(" Delete")
        self.btn_delete.setIcon(qta.icon('fa5s.trash-alt'))
        
        self.btn_import.setEnabled(False)
        self.btn_add.setEnabled(False) 
        self.btn_edit.setEnabled(False) 
        self.btn_delete.setEnabled(False) 
        
        self.btn_add.clicked.connect(self.open_add_dialog)
        self.btn_edit.clicked.connect(self.open_edit_dialog)
        self.btn_delete.clicked.connect(self.delete_selected)
        
        search_layout.addWidget(QLabel("Search In:"))
        search_layout.addWidget(self.col_combo)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.btn_match_case)
        search_layout.addWidget(self.btn_whole_word)
        search_layout.addWidget(self.btn_search)
        search_layout.addWidget(self.btn_import)
        search_layout.addWidget(self.btn_export)
        search_layout.addWidget(self.btn_add)
        search_layout.addWidget(self.btn_edit)
        search_layout.addWidget(self.btn_delete)
        layout.addLayout(search_layout)

        self.table_view = QTableView()
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setSelectionMode(QTableView.SingleSelection)
        self.table_view.setEditTriggers(QTableView.NoEditTriggers)
        self.table_view.doubleClicked.connect(self.on_cell_double_clicked)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_context_menu)
        self.table_view.setSortingEnabled(True)
        
        header = self.table_view.horizontalHeader()
        header.setSectionsMovable(True)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(True)
        header.sortIndicatorChanged.connect(self.on_sort_changed)
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self.show_header_context_menu)
        
        layout.addWidget(self.table_view)

        self.query_model = QSqlQueryModel()
        self.table_view.setModel(self.query_model)
        self.table_view.selectionModel().selectionChanged.connect(self.on_selection_changed)

        page_layout = QHBoxLayout()
        
        self.btn_prev = QPushButton(" Prev") 
        self.btn_prev.setIcon(qta.icon('fa5s.chevron-left'))
        self.btn_prev.clicked.connect(self.prev_page)
        
        self.page_label = QLabel("Page 0 of 0 (Total: 0)")
        self.page_label.setAlignment(Qt.AlignCenter)
        
        self.btn_next = QPushButton("Next ") 
        self.btn_next.setIcon(qta.icon('fa5s.chevron-right'))
        self.btn_next.setLayoutDirection(Qt.RightToLeft)
        self.btn_next.clicked.connect(self.next_page)
        
        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["50", "100", "500", "1000"])
        self.page_size_combo.setCurrentText("100")
        self.page_size_combo.currentTextChanged.connect(self.change_page_size)
        
        page_layout.addStretch()
        page_layout.addWidget(self.btn_prev)
        page_layout.addWidget(self.page_label)
        page_layout.addWidget(self.btn_next)
        page_layout.addWidget(QLabel("  Rows per page:"))
        page_layout.addWidget(self.page_size_combo)
        layout.addLayout(page_layout)

    def load_table_data(self, class_id, class_name):
        self.current_class_id = class_id
        self.current_class_name = class_name
        self.header_label.setText(f"<h2>{class_name}</h2>")
        
        self.btn_import.setEnabled(True)
        self.btn_add.setEnabled(True)
        self.btn_edit.setEnabled(False)
        self.btn_delete.setEnabled(False)
        
        self.db_path = QSettings("MyCompany", "DatabaseManagerApp").value("db_path", "")

        # Release any read lock the Qt view holds on the old class's view before the
        # schema sync runs DDL on this database file (belt-and-suspenders alongside WAL).
        self.query_model.clear()

        try:
            self.current_view_name, self.current_table_name = sync_physical_table(self.db_path, class_id, class_name, parent_widget=self)
            
            if not self.current_view_name:
                raise RuntimeError("Sync aborted or failed to generate a valid view.")
                
            with sqlite3.connect(self.db_path) as conn:
                cur = conn.cursor()
                cur.execute(f"PRAGMA table_info({qid(self.current_view_name)})")
                columns = [row[1] for row in cur.fetchall()]

                cur.execute("SELECT name FROM attributes WHERE class_id = ? AND data_type = 'file'", (class_id,))
                self.file_columns = {row[0]: sanitize_name(row[0]) for row in cur.fetchall()}

            if not columns:
                raise sqlite3.OperationalError(f"View '{self.current_view_name}' is broken and returned no columns.")
                
        except Exception as e:
            QMessageBox.critical(self, "Schema Load Error", f"A database operational error occurred while loading '{class_name}'.\n\nThis is usually caused by broken 'Look-through' attributes referencing deleted or missing columns.\n\nDetails:\n{str(e)}")
            self.header_label.setText("<h2>Sync Aborted</h2>")
            self.btn_import.setEnabled(False)
            self.btn_add.setEnabled(False)
            return
            
        self.search_input.clear()
        self.current_sort_col = "ID"
        self.current_sort_order = "ASC"
        self.current_page = 0
        
        header = self.table_view.horizontalHeader()
        header.blockSignals(True)
        header.setSortIndicator(0, Qt.AscendingOrder)
        header.blockSignals(False)

        self.col_combo.blockSignals(True)
        self.col_combo.clear()
        self.col_combo.addItem("All Columns", -1)
        for i, col_name in enumerate(columns):
            self.col_combo.addItem(col_name, i)
        self.col_combo.blockSignals(False)
        
        self.build_and_exec_query()

    @staticmethod
    def _glob_escape(s):
        # GLOB has no ESCAPE clause; wrap each metachar in a one-char class to match it literally.
        out = []
        for ch in s:
            out.append(f"[{ch}]" if ch in "*?[" else ch)
        return "".join(out)

    @staticmethod
    def _like_escape(s):
        return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")

    def _build_text_condition(self, col_sql, search_text):
        """Build a (sql_condition, params) pair honouring the Match Case / Whole Word toggles.

        The same expression is used by both the COUNT and the data query, so paging stays
        consistent. Works on both the sqlite3 and the Qt connection (it's plain SQLite SQL).
        Case-insensitive matching is ASCII-only, the same limitation LIKE already had.
        """
        match_case = self.btn_match_case.isChecked()
        whole_word = self.btn_whole_word.isChecked()
        col_expr = f"CAST({col_sql} AS TEXT)"

        if whole_word:
            boundary = "[^A-Za-z0-9_]"
            term = self._glob_escape(search_text)
            if not match_case:
                col_expr = f"LOWER({col_expr})"
                term = term.lower()
            patterns = [
                term,                                      # whole value equals the word
                f"{term}{boundary}*",                      # word at the start
                f"*{boundary}{term}",                      # word at the end
                f"*{boundary}{term}{boundary}*",           # word in the middle
            ]
            cond = "(" + " OR ".join([f"{col_expr} GLOB ?"] * len(patterns)) + ")"
            return cond, patterns

        if match_case:
            # INSTR is case- and accent-sensitive (binary), i.e. true "match case".
            return f"INSTR({col_expr}, ?) > 0", [search_text]

        esc = self._like_escape(search_text)
        return f"{col_expr} LIKE ? ESCAPE '\\'", [f"%{esc}%"]

    def _build_search_where(self, cur):
        """Build the WHERE clause for the current search box, honouring the toggles.

        Returns (where_sql, params); where_sql is '' or starts with ' WHERE '.
        Shared by the table view and 'current filter only' exports so they agree.
        """
        where_clauses = []
        params = []
        search_text = self.search_input.text().strip()
        if search_text:
            col_idx = self.col_combo.currentData()
            if col_idx == -1:
                cur.execute(f"PRAGMA table_info({qid(self.current_view_name)})")
                col_names = [row[1] for row in cur.fetchall()]
                or_clauses = []
                for c in col_names:
                    cond, cond_params = self._build_text_condition(qid(c), search_text)
                    or_clauses.append(cond)
                    params.extend(cond_params)
                if or_clauses:
                    where_clauses.append("(" + " OR ".join(or_clauses) + ")")
            else:
                col_name = self.col_combo.itemText(self.col_combo.currentIndex())
                cond, cond_params = self._build_text_condition(qid(col_name), search_text)
                where_clauses.append(cond)
                params.extend(cond_params)
        where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
        return where_sql, params

    def _get_import_schema(self, cur):
        """The canonical, round-trippable column schema for this class.

        Returns (attributes, relationships, lookthrough_names) where the import-ready
        header order is ['ID'] + [a.name for attributes] + [r.name for relationships].
        """
        cur.execute("SELECT id, name, data_type, is_unique, is_required, lookup_query FROM attributes WHERE class_id = ? ORDER BY row_order", (self.current_class_id,))
        attr_rows = cur.fetchall()

        cur.execute("""
            SELECT a.name, COUNT(m.id) FROM attributes a
            JOIN matrix_columns m ON a.id = m.attribute_id
            WHERE a.class_id = ? GROUP BY a.id
        """, (self.current_class_id,))
        matrix_counts = {r[0]: r[1] for r in cur.fetchall()}

        attributes = []
        lookthrough_names = []
        for aid, name, dtype, uniq, req, lookup_query in attr_rows:
            if dtype == "look-through":
                lookthrough_names.append(name)
                continue
            options = None
            if dtype == "discrete":
                try:
                    options = {r[0] for r in cur.execute(
                        "SELECT value FROM discrete_options WHERE type_id = ?", (int(lookup_query),)).fetchall()}
                except (ValueError, TypeError, sqlite3.OperationalError):
                    options = set()
            attributes.append({
                "name": name, "safe": sanitize_name(name), "type": dtype,
                "unique": bool(uniq), "required": bool(req),
                "matrix_count": matrix_counts.get(name, 0),
                "options": options,
            })

        cur.execute("""
            SELECT c.id, c.name FROM relationships r
            JOIN classes c ON r.target_class = c.id
            WHERE r.source_class = ? ORDER BY r.row_order
        """, (self.current_class_id,))
        relationships = []
        seen = set()
        for tcid, tname in cur.fetchall():
            safe_t = sanitize_name(tname)
            if safe_t in seen:
                continue
            seen.add(safe_t)
            cur.execute("SELECT name FROM attributes WHERE class_id = ? AND is_title = 1 ORDER BY row_order", (tcid,))
            title_cols = [r[0] for r in cur.fetchall()]
            relationships.append({
                "name": tname, "safe": safe_t,
                "target_table": f"objects_{safe_t}",
                "junc_table": f"rel_{self.current_table_name}_to_objects_{safe_t}",
                "base_view": f"base_view_objects_{safe_t}",
                "title_cols": title_cols,
            })
        return attributes, relationships, lookthrough_names

    @staticmethod
    def _cell_to_value(cell):
        """Normalise a spreadsheet/CSV cell to a trimmed string or None."""
        if isinstance(cell, datetime.datetime):
            return cell.strftime("%Y-%m-%d %H:%M:%S")
        if cell is None:
            return None
        s = str(cell).strip()
        return s if s != "" else None

    def build_and_exec_query(self):
        if not self.current_view_name: return
        
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()

            base_query = f"FROM {qid(self.current_view_name)}"
            where_sql, params = self._build_search_where(cur)

            count_query = f"SELECT COUNT(*) {base_query} {where_sql}"
            cur.execute(count_query, params)
            self.total_records = cur.fetchone()[0]
            self.total_pages = max(1, (self.total_records + self.page_size - 1) // self.page_size)
            
            if self.current_page >= self.total_pages:
                self.current_page = max(0, self.total_pages - 1)
                
            offset = self.current_page * self.page_size
            order_sql = f"ORDER BY {qid(self.current_sort_col)} {self.current_sort_order}"
            final_query = f"SELECT * {base_query} {where_sql} {order_sql} LIMIT {self.page_size} OFFSET {offset}"
            
            db = QSqlDatabase.database()
            query = QSqlQuery(db)
            query.prepare(final_query)
            for p in params: query.addBindValue(p)
            
            query.exec()
            if query.lastError().isValid():
                QMessageBox.critical(self, "Database Query Error", f"An error occurred while fetching data from the view:\n\n{query.lastError().text()}")
                
            self.query_model.setQuery(query)
            
            hidden_for_class = self.hidden_columns_memory.get(self.current_class_name, set())
            for i in range(self.query_model.columnCount()):
                col_name = self.query_model.headerData(i, Qt.Horizontal)
                if col_name in hidden_for_class:
                    self.table_view.setColumnHidden(i, True)
                else:
                    self.table_view.setColumnHidden(i, False)
            
            self.page_label.setText(f"Page {self.current_page + 1} of {self.total_pages} (Total: {self.total_records})")
            self.btn_prev.setEnabled(self.current_page > 0)
            self.btn_next.setEnabled(self.current_page < self.total_pages - 1)

    def on_sort_changed(self, logical_index, order):
        col_name = self.query_model.headerData(logical_index, Qt.Horizontal)
        if not col_name: return 
        self.current_sort_col = col_name
        self.current_sort_order = "ASC" if order == Qt.AscendingOrder else "DESC"
        self.build_and_exec_query()

    def trigger_search(self):
        self.current_page = 0
        self.build_and_exec_query()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.build_and_exec_query()
            
    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.build_and_exec_query()
            
    def change_page_size(self, new_size):
        self.page_size = int(new_size)
        self.current_page = 0
        self.build_and_exec_query()

    def on_selection_changed(self):
        has_sel = bool(self.table_view.selectionModel().selectedRows())
        self.btn_edit.setEnabled(has_sel)
        self.btn_delete.setEnabled(has_sel)

    def get_selected_id(self):
        rows = self.table_view.selectionModel().selectedRows()
        if not rows: return None
        idx = rows[0]
        id_index = self.query_model.index(idx.row(), 0)
        return self.query_model.data(id_index)

    def show_header_context_menu(self, position):
        header = self.table_view.horizontalHeader()
        logical_index = header.logicalIndexAt(position)
        menu = QMenu()
        
        act_hide = None
        if logical_index >= 0:
            col_name = self.query_model.headerData(logical_index, Qt.Horizontal)
            if col_name:
                visible_count = sum(1 for i in range(header.count()) if not self.table_view.isColumnHidden(i))
                if visible_count > 1: act_hide = menu.addAction(f"Hide '{col_name}' Temporarily")
        
        hidden_cols = [i for i in range(header.count()) if self.table_view.isColumnHidden(i)]
        act_unhide = None
        if hidden_cols:
            if act_hide: menu.addSeparator()
            act_unhide = menu.addAction("Unhide All Columns")
            
        if not menu.actions(): return
            
        action = menu.exec(header.viewport().mapToGlobal(position))
        
        if act_hide and action == act_hide:
            self.table_view.setColumnHidden(logical_index, True)
            col_name = self.query_model.headerData(logical_index, Qt.Horizontal)
            if col_name: self.hidden_columns_memory.setdefault(self.current_class_name, set()).add(col_name)
        elif act_unhide and action == act_unhide:
            for i in range(header.count()): self.table_view.setColumnHidden(i, False)
            if self.current_class_name in self.hidden_columns_memory:
                self.hidden_columns_memory[self.current_class_name].clear()

    def show_context_menu(self, position):
        if not self.table_view.selectionModel().selectedRows(): return
        from PySide6.QtWidgets import QMenu
        menu = QMenu()
        act_edit = menu.addAction(qta.icon('fa5s.edit'), "Edit Object")
        act_del = menu.addAction(qta.icon('fa5s.trash-alt'), "Delete Object")
        
        action = menu.exec(self.table_view.viewport().mapToGlobal(position))
        if action == act_edit: self.open_edit_dialog()
        elif action == act_del: self.delete_selected()

    def open_add_dialog(self):
        dialog = ObjectEditorDialog(self.db_path, self.current_class_id, self.current_class_name, self.current_table_name, None, self)
        if dialog.exec(): self.build_and_exec_query() 
            
    def on_cell_double_clicked(self, index):
        # Double-clicking a 'file' cell opens/saves the attachment; any other cell edits the row.
        header = self.query_model.headerData(index.column(), Qt.Horizontal)
        if header in self.file_columns:
            self._open_file_cell(index, self.file_columns[header])
        else:
            self.open_edit_dialog()

    def _open_file_cell(self, index, safe_col):
        id_index = self.query_model.index(index.row(), 0)
        obj_id = self.query_model.data(id_index)
        if not obj_id:
            return
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(f"SELECT {qid(safe_col)} FROM {qid(self.current_table_name)} WHERE id = ?", (obj_id,)).fetchone()
        stored = row[0] if row else None
        if not stored:
            QMessageBox.information(self, "No File", "There is no file attached to this cell.")
            return

        path = resolve_file_path(self.db_path, stored)
        display = display_file_name(stored)
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "File Not Found",
                f"'{display}' is referenced by this record but is no longer in the files folder.")
            return

        box = QMessageBox(self)
        box.setWindowTitle("File")
        box.setIcon(QMessageBox.Question)
        box.setText(f"<b>{display}</b>")
        box.setInformativeText("What would you like to do?")
        btn_open = box.addButton(" Open", QMessageBox.AcceptRole)
        btn_open.setIcon(qta.icon('fa5s.external-link-alt'))
        btn_save = box.addButton(" Save As...", QMessageBox.ActionRole)
        btn_save.setIcon(qta.icon('fa5s.download'))
        box.addButton(QMessageBox.Cancel)
        box.exec()

        clicked = box.clickedButton()
        if clicked == btn_open:
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))
        elif clicked == btn_save:
            dest, _ = QFileDialog.getSaveFileName(self, "Save File As", display, "All Files (*.*)")
            if dest:
                try:
                    shutil.copy2(path, dest)
                except OSError as e:
                    QMessageBox.critical(self, "Save Failed", str(e))

    def open_edit_dialog(self):
        obj_id = self.get_selected_id()
        if not obj_id: return
        dialog = ObjectEditorDialog(self.db_path, self.current_class_id, self.current_class_name, self.current_table_name, obj_id, self)
        if dialog.exec(): self.build_and_exec_query()

    def delete_selected(self):
        obj_id = self.get_selected_id()
        if not obj_id: return
        reply = QMessageBox.question(self, "Confirm Delete", f"Are you sure you want to delete this {self.current_class_name}?\n\nIts relationships will be automatically safely removed.", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                # Collect attached files first so we can retire them after a successful delete.
                stored_files = []
                if self.file_columns:
                    with sqlite3.connect(self.db_path) as conn:
                        cols = list(self.file_columns.values())
                        sel = ", ".join(qid(c) for c in cols)
                        r = conn.execute(f"SELECT {sel} FROM {qid(self.current_table_name)} WHERE id = ?", (obj_id,)).fetchone()
                        if r:
                            stored_files = [v for v in r if v]

                with sqlite3.connect(self.db_path) as conn:
                    conn.execute("PRAGMA foreign_keys = 1")
                    conn.execute("BEGIN IMMEDIATE")
                    conn.execute(f"DELETE FROM {qid(self.current_table_name)} WHERE id = ?", (obj_id,))
                    conn.commit()

                for stored in stored_files:
                    trash_stored_file(self.db_path, stored)

                self.build_and_exec_query()
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def export_data(self):
        if not self.current_view_name: return

        has_filter = bool(self.search_input.text().strip())
        opt_dlg = ExportOptionsDialog(self, has_filter=has_filter)
        if not opt_dlg.exec(): return
        opts = opt_dlg.get_options()

        ext = ".xlsx" if opts['fmt'] == 'xlsx' else ".csv"
        filt = "Excel Files (*.xlsx)" if opts['fmt'] == 'xlsx' else "CSV Files (*.csv)"
        default_name = f"{sanitize_name(self.current_class_name)}{'_template' if opts['template'] else ''}{ext}"
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Data", default_name, filt)
        if not file_path: return
        if not file_path.lower().endswith(ext): file_path += ext

        try:
            headers, rows = self._gather_export_rows(opts)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not gather data to export:\n\n{e}")
            return

        try:
            if opts['fmt'] == 'xlsx':
                try:
                    import openpyxl
                except ImportError:
                    file_path = file_path[:-5] + ".csv" if file_path.lower().endswith(".xlsx") else file_path + ".csv"
                    self._write_csv(file_path, headers, rows)
                    QMessageBox.warning(self, "Exported as CSV",
                        f"'openpyxl' is not installed, so the data was exported as CSV instead:\n{file_path}\n\n"
                        "Install it (pip install openpyxl) for Excel export.")
                    return
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
                for row in rows: ws.append(row)
                wb.save(file_path)
            else:
                self._write_csv(file_path, headers, rows)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to write file:\n\n{e}")
            return

        what = "template" if opts['template'] else f"{len(rows)} row(s)"
        QMessageBox.information(self, "Export Complete", f"Exported {what} to:\n{file_path}")

    @staticmethod
    def _write_csv(file_path, headers, rows):
        with open(file_path, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    def _gather_export_rows(self, opts):
        """Return (headers, rows) for the chosen export options."""
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()

            if opts['import_ready']:
                attrs, rels, _lookthroughs = self._get_import_schema(cur)
                headers = ["ID"] + [a['name'] for a in attrs] + [r['name'] for r in rels]
                if opts['template']:
                    return headers, []

                selects = ["m.id"]
                # 'file' columns export the display name (not the raw "<uuid>__name" reference).
                selects += [
                    (f"substr(m.{qid(a['safe'])}, {FILE_PREFIX_LEN + 1})" if a['type'] == 'file' else f"m.{qid(a['safe'])}")
                    for a in attrs
                ]
                selects += [f"(SELECT GROUP_CONCAT(target_id) FROM {qid(r['junc_table'])} WHERE source_id = m.id)" for r in rels]
                sql = f"SELECT {', '.join(selects)} FROM {qid(self.current_table_name)} m"

                params = []
                if opts['rows_filtered']:
                    where_sql, wparams = self._build_search_where(cur)
                    cur.execute(f"SELECT [ID] FROM {qid(self.current_view_name)}{where_sql}", wparams)
                    ids = [r[0] for r in cur.fetchall()]
                    if not ids:
                        return headers, []
                    sql += f" WHERE m.id IN ({', '.join(['?'] * len(ids))})"
                    params = ids
                sql += " ORDER BY m.id ASC"
                cur.execute(sql, params)
                raw = cur.fetchall()
            else:
                cur.execute(f"PRAGMA table_info({qid(self.current_view_name)})")
                headers = [row[1] for row in cur.fetchall()]
                if opts['template']:
                    return headers, []
                where_sql, params = self._build_search_where(cur) if opts['rows_filtered'] else ("", [])
                order_sql = f"ORDER BY {qid(self.current_sort_col)} {self.current_sort_order}"
                cur.execute(f"SELECT * FROM {qid(self.current_view_name)}{where_sql} {order_sql}", params)
                raw = cur.fetchall()

        return headers, [["" if v is None else str(v) for v in row] for row in raw]

    def _read_table_file(self, file_path):
        """Read an .xlsx or .csv file into a list of row-lists (header row first)."""
        if file_path.lower().endswith(".csv"):
            with open(file_path, newline='', encoding='utf-8-sig') as f:
                return [list(r) for r in csv.reader(f)]
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("The 'openpyxl' library is required to read .xlsx files.\nInstall it with: pip install openpyxl")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _analyze_import(self, all_rows, attrs, rels, lookthrough_names):
        """Classify each column and decide whether the file can be imported."""
        attr_by_safe = {a['safe']: a for a in attrs}
        rel_by_safe = {r['safe']: r for r in rels}
        look_safe = {sanitize_name(n) for n in lookthrough_names}

        headers_raw = all_rows[0] if all_rows else []
        col_map = []      # one entry per column: (kind, obj)
        columns = []      # (display, label, kind) for the preview
        seen = set()
        duplicates = []
        id_col_index = None
        mapped_attr_safes = set()

        for idx, h in enumerate(headers_raw):
            disp = "" if h is None else str(h).strip()
            s = sanitize_name(disp) if disp else ""
            if not disp:
                col_map.append(('blank', None)); continue
            if s == "id":
                id_col_index = idx
                col_map.append(('id', None))
                columns.append((disp, "ID (row key for update)", "key")); continue
            if s in attr_by_safe:
                if s in seen: duplicates.append(disp)
                seen.add(s); mapped_attr_safes.add(s)
                a = attr_by_safe[s]
                col_map.append(('attr', a))
                label = "File (loaded from this sheet's folder)" if a['type'] == 'file' else "Attribute"
                columns.append((disp, label, "ok")); continue
            if s in rel_by_safe:
                if s in seen: duplicates.append(disp)
                seen.add(s)
                col_map.append(('rel', rel_by_safe[s]))
                columns.append((disp, "Relationship", "ok")); continue
            if s in look_safe:
                col_map.append(('look', None))
                columns.append((disp, "Look-through (computed, ignored)", "ignore")); continue
            col_map.append(('unknown', None))
            columns.append((disp, "Unrecognized (ignored)", "ignore"))

        data_count = sum(1 for r in all_rows[1:] if any(c is not None and str(c).strip() != "" for c in r))
        missing_required = [a['name'] for a in attrs if a['required'] and a['safe'] not in mapped_attr_safes]
        has_recognized = any(k in ('attr', 'rel') for k, _ in col_map)

        problems = []
        if duplicates:
            problems.append("Duplicate columns map to the same field: " + ", ".join(duplicates))
        if missing_required:
            problems.append("Required columns are missing: " + ", ".join(missing_required))
        if data_count == 0:
            problems.append("No data rows found.")
        if not has_recognized:
            problems.append("No recognized attribute or relationship columns.")

        return {
            'columns': columns, 'col_map': col_map, 'id_col_index': id_col_index,
            'data_count': data_count, 'missing_required': missing_required,
            'duplicates': duplicates, 'problems': problems, 'can_import': not problems,
        }

    def import_data(self):
        if not self.current_view_name: return

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Data", "",
            "Data Files (*.xlsx *.csv);;Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not file_path: return

        try:
            all_rows = self._read_table_file(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", f"Could not read the file:\n\n{e}")
            return

        if not all_rows:
            QMessageBox.warning(self, "Import", "The file appears to be empty.")
            return

        with sqlite3.connect(self.db_path) as conn:
            attrs, rels, lookthrough_names = self._get_import_schema(conn.cursor())

        analysis = self._analyze_import(all_rows, attrs, rels, lookthrough_names)
        analysis['file'] = os.path.basename(file_path)

        dlg = ImportPreviewDialog(self, analysis)
        if not dlg.exec(): return

        self._run_import(all_rows, attrs, rels, analysis, os.path.dirname(os.path.abspath(file_path)))

    def _run_import(self, all_rows, attrs, rels, analysis, source_dir):
        col_map = analysis['col_map']
        id_idx = analysis['id_col_index']
        table = self.current_table_name

        copied_abs = []     # files copied into storage this run (deleted on failure)
        old_to_trash = []   # replaced file references to retire after a successful commit
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("PRAGMA foreign_keys = 1")
                cur = conn.cursor()

                # --- snapshots for validation ---
                cur.execute(f"SELECT id FROM {qid(table)}")
                existing_ids = {r[0] for r in cur.fetchall()}

                uniq_maps = {}  # safe_col -> {value: owner_id}
                for a in attrs:
                    if a['unique']:
                        cur.execute(f"SELECT {qid(a['safe'])}, id FROM {qid(table)} WHERE {qid(a['safe'])} IS NOT NULL")
                        uniq_maps[a['safe']] = {v: i for v, i in cur.fetchall()}

                rel_data = {}  # safe -> {'ids': set, 'title_map': {lower: [ids]} or None, 'junc': str, 'name': str}
                for r in rels:
                    cur.execute(f"SELECT id FROM {qid(r['target_table'])}")
                    tids = {x[0] for x in cur.fetchall()}
                    title_map = None
                    if len(r['title_cols']) == 1:
                        title_map = {}
                        try:
                            cur.execute(f"SELECT [ID], {qid(r['title_cols'][0])} FROM {qid(r['base_view'])}")
                            for tid, tval in cur.fetchall():
                                if tval is not None and str(tval).strip() != "":
                                    title_map.setdefault(str(tval).strip().lower(), []).append(tid)
                        except sqlite3.OperationalError:
                            title_map = None
                    rel_data[r['safe']] = {'ids': tids, 'title_map': title_map,
                                           'junc': r['junc_table'], 'name': r['name']}

                # --- validate every row first (nothing written yet) ---
                file_unique_used = {}  # (safe, value) -> row_number
                seen_ids = set()
                ops = []

                for rn, row in enumerate(all_rows[1:], start=2):
                    if not any(c is not None and str(c).strip() != "" for c in row):
                        continue

                    mode, target_id = 'insert', None
                    if id_idx is not None and id_idx < len(row):
                        idraw = row[id_idx]
                        if idraw is not None and str(idraw).strip() != "":
                            try:
                                target_id = int(float(str(idraw).strip()))
                            except (ValueError, TypeError):
                                raise ValueError(f"Row {rn}: invalid ID '{idraw}'.")
                            if target_id in seen_ids:
                                raise ValueError(f"Row {rn}: ID {target_id} appears more than once in the file.")
                            seen_ids.add(target_id)
                            if target_id in existing_ids:
                                mode = 'update'
                            else:
                                mode, target_id = 'insert', None  # unknown ID -> add as new

                    set_cols, set_vals, rels_in_row, file_ops_row = [], [], {}, []

                    for idx, (kind, obj) in enumerate(col_map):
                        if idx >= len(row):
                            break
                        if kind == 'attr' and obj['type'] == 'file':
                            a = obj
                            val = self._cell_to_value(row[idx])
                            if val is None:
                                # Empty cell = keep the existing attachment (do not clear).
                                if a['required'] and mode == 'insert':
                                    raise ValueError(f"Row {rn}: '{a['name']}' is required but empty.")
                                continue
                            fname = os.path.basename(val)
                            src = os.path.join(source_dir, fname)
                            if not os.path.isfile(src):
                                raise ValueError(f"Row {rn}: file '{fname}' for '{a['name']}' was not found next to the import file.")
                            new_stored = make_stored_filename(fname)
                            set_cols.append(a['safe'])
                            set_vals.append(new_stored)
                            file_ops_row.append((a['safe'], src, new_stored))
                            continue
                        if kind == 'attr':
                            a = obj
                            val = self._cell_to_value(row[idx])
                            if val is None:
                                if a['required']:
                                    raise ValueError(f"Row {rn}: '{a['name']}' is required but empty.")
                                val_final = None
                            else:
                                ok, val_final = safe_convert(val, a['type'])
                                if not ok:
                                    raise ValueError(f"Row {rn}: cannot convert '{val}' for '{a['name']}' ({a['type']}).")
                                if a['type'] == 'discrete' and a['options'] is not None and val_final not in a['options']:
                                    raise ValueError(f"Row {rn}: '{val_final}' is not a valid option for '{a['name']}'.")
                                if a['type'] == 'matrix' and a['matrix_count']:
                                    try:
                                        parsed = ast.literal_eval(val_final)
                                    except (ValueError, SyntaxError):
                                        parsed = None
                                    if not isinstance(parsed, list) or len(parsed) != a['matrix_count']:
                                        raise ValueError(f"Row {rn}: '{a['name']}' must be a list of exactly {a['matrix_count']} column list(s).")
                                if a['unique'] and val_final is not None:
                                    owner = uniq_maps.get(a['safe'], {}).get(val_final)
                                    if owner is not None and owner != target_id:
                                        raise ValueError(f"Row {rn}: '{a['name']}' value '{val_final}' already exists in the database.")
                                    key = (a['safe'], val_final)
                                    if key in file_unique_used:
                                        raise ValueError(f"Row {rn}: '{a['name']}' value '{val_final}' is duplicated in the file (also row {file_unique_used[key]}).")
                                    file_unique_used[key] = rn
                            set_cols.append(a['safe'])
                            set_vals.append(val_final)

                        elif kind == 'rel':
                            r = obj
                            val = self._cell_to_value(row[idx])
                            rd = rel_data[r['safe']]
                            tids = []
                            if val is not None:
                                for tok in val.split(','):
                                    tok = tok.strip()
                                    if tok == "":
                                        continue
                                    if re.match(r'^-?\d+$', tok):
                                        tid = int(tok)
                                        if tid not in rd['ids']:
                                            raise ValueError(f"Row {rn}: '{r['name']}' target ID {tid} does not exist.")
                                    elif rd['title_map'] is not None:
                                        matches = rd['title_map'].get(tok.lower())
                                        if not matches:
                                            raise ValueError(f"Row {rn}: '{r['name']}' has no object titled '{tok}'.")
                                        if len(matches) > 1:
                                            raise ValueError(f"Row {rn}: '{r['name']}' title '{tok}' is ambiguous ({len(matches)} matches); use a numeric ID.")
                                        tid = matches[0]
                                    else:
                                        raise ValueError(f"Row {rn}: '{r['name']}' must use numeric IDs (target has no single title column).")
                                    tids.append(tid)
                            rels_in_row[r['safe']] = tids  # column present -> replace links

                    ops.append({'mode': mode, 'id': target_id, 'cols': set_cols, 'vals': set_vals,
                                'rels': rels_in_row, 'files': file_ops_row})

                # --- apply atomically ---
                inserted = updated = 0
                files_dir = files_dir_for(self.db_path, create=True) if any(op['files'] for op in ops) else None
                conn.execute("BEGIN IMMEDIATE")
                for op in ops:
                    # On update, remember the file(s) we're about to replace so we can trash them.
                    if op['mode'] == 'update' and op['files']:
                        cols = [f[0] for f in op['files']]
                        sel = ", ".join(qid(c) for c in cols)
                        oldrow = cur.execute(f"SELECT {sel} FROM {qid(table)} WHERE id = ?", (op['id'],)).fetchone()
                        if oldrow:
                            old_to_trash.extend([v for v in oldrow if v])

                    # Copy new attachments into storage before the row write.
                    for (safe_col, src, new_stored) in op['files']:
                        dest_abs = os.path.join(files_dir, new_stored)
                        shutil.copy2(src, dest_abs)
                        copied_abs.append(dest_abs)

                    if op['mode'] == 'update':
                        oid = op['id']
                        if op['cols']:
                            set_clause = ", ".join(f"{qid(c)} = ?" for c in op['cols'])
                            cur.execute(f"UPDATE {qid(table)} SET {set_clause} WHERE id = ?", op['vals'] + [oid])
                        updated += 1
                    else:
                        if op['cols']:
                            ph = ", ".join(["?"] * len(op['cols']))
                            cur.execute(f"INSERT INTO {qid(table)} ({', '.join(qid(c) for c in op['cols'])}) VALUES ({ph})", op['vals'])
                        else:
                            cur.execute(f"INSERT INTO {qid(table)} DEFAULT VALUES")
                        oid = cur.lastrowid
                        inserted += 1

                    for rsafe, tids in op['rels'].items():
                        junc = rel_data[rsafe]['junc']
                        cur.execute(f"DELETE FROM {qid(junc)} WHERE source_id = ?", (oid,))
                        for tid in tids:
                            cur.execute(f"INSERT INTO {qid(junc)} (source_id, target_id) VALUES (?, ?)", (oid, tid))

                conn.commit()

            for old in old_to_trash:
                trash_stored_file(self.db_path, old)

            self.build_and_exec_query()
            QMessageBox.information(self, "Import Complete",
                f"Added {inserted} and updated {updated} object(s).")
        except Exception as e:
            for path in copied_abs:
                try:
                    os.remove(path)
                except OSError:
                    pass
            QMessageBox.critical(self, "Import Failed",
                f"Import aborted. No changes were made to the database.\n\nReason:\n{e}")


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Nexus")
        self.setWindowIcon(get_app_icon())
        self.resize(1200, 768)

        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)

        self.setup_menu()

        splitter = QSplitter(Qt.Horizontal)
        self.main_layout.addWidget(splitter)

        # ---------------------------------------------
        # LEFT PANEL: Splitting Classes and Modules
        # ---------------------------------------------
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(9, 9, 9, 9)
        
        # Dummy header to perfectly align the top of the sidebars with the search layout
        dummy_header = QLabel("<h2>&nbsp;</h2>")
        left_layout.addWidget(dummy_header)
        
        left_splitter = QSplitter(Qt.Vertical)
        
        # 1. Classes Tree (Headers removed)
        self.sidebar = QTreeWidget()
        self.sidebar.setHeaderHidden(True)
        self.sidebar.itemClicked.connect(self.change_class)
        left_splitter.addWidget(self.sidebar)

        # 2. Modules Tree (Headers removed)
        self.module_sidebar = QTreeWidget()
        self.module_sidebar.setHeaderHidden(True)
        self.module_sidebar.setContextMenuPolicy(Qt.CustomContextMenu)
        self.module_sidebar.customContextMenuRequested.connect(self.show_module_context_menu)
        left_splitter.addWidget(self.module_sidebar)

        # Allocate 75% height to classes, 25% to modules
        left_splitter.setSizes([750, 250])
        left_layout.addWidget(left_splitter)
        
        # Dummy footer to perfectly align the bottom of the sidebars with the table view bottom
        dummy_footer_layout = QHBoxLayout()
        dummy_btn = QPushButton(" ")
        dummy_btn.setFlat(True)
        dummy_btn.setEnabled(False)
        dummy_btn.setStyleSheet("background: transparent; color: transparent; border: none;")
        dummy_footer_layout.addWidget(dummy_btn)
        left_layout.addLayout(dummy_footer_layout)
        
        splitter.addWidget(left_widget)

        # ---------------------------------------------
        # RIGHT PANEL: Data Browser
        # ---------------------------------------------
        self.data_browser = DataBrowserPage(self)
        splitter.addWidget(self.data_browser)
        
        splitter.setSizes([250, 950])
        self.refresh_sidebar()

    def setup_menu(self):
        menubar = QMenuBar(self)
        self.main_layout.setMenuBar(menubar)
        
        db_menu = menubar.addMenu("Database")

        action_settings = QAction(qta.icon('fa5s.cog'), " Settings", self)
        action_settings.triggered.connect(self.open_settings)
        db_menu.addAction(action_settings)

        action_builder = QAction(qta.icon('fa5s.tools'), " Class Builder", self)
        action_builder.triggered.connect(self.open_builder)
        db_menu.addAction(action_builder)

        action_discrete = QAction(qta.icon('fa5s.list-ul'), " Discrete Type Builder", self)
        action_discrete.triggered.connect(self.open_discrete_builder)
        db_menu.addAction(action_discrete)

        action_module = QAction(qta.icon('fa5s.file-code'), " Module Builder", self)
        action_module.triggered.connect(self.open_module_builder)
        db_menu.addAction(action_module)

        db_menu.addSeparator()
        action_cleanup = QAction(qta.icon('fa5s.broom'), " Clean Up Unused Files", self)
        action_cleanup.triggered.connect(self.clean_unused_files)
        db_menu.addAction(action_cleanup)

        doc_menu = menubar.addMenu("Documentation")
        action_api = QAction(qta.icon('fa5s.book'), " Module API", self)
        action_api.triggered.connect(self.open_module_api)
        doc_menu.addAction(action_api)

    def open_settings(self):
        dialog = SettingsDialog(self)
        dialog.exec()
        self.refresh_sidebar()

    def open_builder(self):
        dialog = ClassBuilderDialog(self)
        dialog.exec()
        self.refresh_sidebar()

    def open_discrete_builder(self):
        dialog = DiscreteTypeBuilderDialog(self)
        dialog.exec()

    def open_module_builder(self):
        db_path = QSettings("MyCompany", "DatabaseManagerApp").value("db_path", "")
        dialog = ModuleBuilderDialog(db_path, self)
        dialog.exec()
        self.refresh_module_sidebar()
        
    def open_module_api(self):
        QMessageBox.information(self, "Module API", "Documentation coming soon...\n\nCurrently available built-in functions:\n\nget_objects(class_name)\n-> Returns a complete Pandas DataFrame of the specified class, including automatically resolved multi-title relationships.")

    def clean_unused_files(self):
        """Move any file in the storage folder that no record references into _trash."""
        db_path = QSettings("MyCompany", "DatabaseManagerApp").value("db_path", "")
        files_dir = files_dir_for(db_path, create=False)
        if not files_dir or not os.path.isdir(files_dir):
            QMessageBox.information(self, "Clean Up", "There is no files folder yet — nothing to clean.")
            return

        referenced = set()
        try:
            with sqlite3.connect(db_path) as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT c.name, a.name FROM attributes a
                    JOIN classes c ON a.class_id = c.id
                    WHERE a.data_type = 'file'
                """)
                file_attrs = cur.fetchall()
                for class_name, attr_name in file_attrs:
                    table = f"objects_{sanitize_name(class_name)}"
                    col = sanitize_name(attr_name)
                    try:
                        cur.execute(f"SELECT {qid(col)} FROM {qid(table)} WHERE {qid(col)} IS NOT NULL")
                        referenced.update(r[0] for r in cur.fetchall() if r[0])
                    except sqlite3.OperationalError:
                        continue
        except sqlite3.OperationalError:
            pass

        moved = 0
        for entry in os.listdir(files_dir):
            full = os.path.join(files_dir, entry)
            if entry == "_trash" or not os.path.isfile(full):
                continue
            if entry not in referenced:
                trash_stored_file(db_path, entry)
                moved += 1

        QMessageBox.information(self, "Clean Up Complete",
            f"Moved {moved} unused file(s) to the _trash folder inside:\n{files_dir}")

    def sync_all_classes(self, db_path):
        with sqlite3.connect(db_path) as conn:
            cur = conn.cursor()
            
            try:
                cur.execute("SELECT id, name FROM classes")
                classes = {row[0]: row[1] for row in cur.fetchall()}
            except sqlite3.OperationalError:
                classes = {}
                
            dependencies = {cid: set() for cid in classes}
            try:
                cur.execute("SELECT class_id, lookup_query FROM attributes WHERE data_type = 'look-through' AND lookup_query != ''")
                for cid, lookup in cur.fetchall():
                    if lookup:
                        tgt_class_name = lookup.split('.')[0].strip().lower()
                        for t_id, t_name in classes.items():
                            if t_name.lower() == tgt_class_name: dependencies[cid].add(t_id)
            except sqlite3.OperationalError:
                pass 
                
        ordered_cids = []
        visited = set()
        temp_mark = set()
        cycle_detected = False
        
        def visit(n):
            if n in temp_mark: return False 
            if n not in visited:
                temp_mark.add(n)
                for m in dependencies.get(n, set()):
                    if m in classes and not visit(m): return False
                temp_mark.remove(n)
                visited.add(n)
                ordered_cids.append(n)
            return True
            
        for cid in classes:
            if cid not in visited:
                if not visit(cid):
                    cycle_detected = True
                    break
                    
        if cycle_detected: ordered_cids = list(classes.keys())
        
        sync_errors = []
        for cid in ordered_cids:
            try:
                sync_physical_table(db_path, cid, classes[cid], parent_widget=self)
            except Exception as e:
                sync_errors.append(f"Class '{classes[cid]}': {str(e)}")
                
        if sync_errors:
            error_msg = "The following configuration errors were detected during sync. Some views might not load perfectly until you fix their schemas:\n\n" + "\n".join(sync_errors)
            QMessageBox.warning(self, "Database Schema Warnings", error_msg)

    def refresh_sidebar(self):
        self.sidebar.clear()
        settings = QSettings("MyCompany", "DatabaseManagerApp")
        db_path = settings.value("db_path", "")
        
        if not db_path:
            db_path = os.path.join(os.path.expanduser("~"), "nexus_default.db")
            settings.setValue("db_path", db_path)
            
        if not os.path.exists(db_path):
            init_db(db_path)
        else:
            # Ensure WAL is enabled even for databases created before this was added.
            # WAL is a persistent property of the file, so this one-time switch sticks.
            try:
                with sqlite3.connect(db_path) as conn:
                    conn.execute("PRAGMA journal_mode=WAL")
            except sqlite3.Error:
                pass

        if QSqlDatabase.contains(): db = QSqlDatabase.database()
        else: db = QSqlDatabase.addDatabase("QSQLITE")

        db.setDatabaseName(db_path)
        if not db.open(): return

        # Drop any read lock the data view holds before sync_all_classes runs DDL.
        self.data_browser.query_model.clear()
        self.sync_all_classes(db_path)

        try:
            with sqlite3.connect(db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT id, name, path FROM classes ORDER BY path ASC, name ASC")
                root_items = {}
                parent_item = self.sidebar.invisibleRootItem()
                
                for c_id, c_name, c_path in cur.fetchall():
                    c_path = (c_path or "").strip().strip('/')
                    current_parent = parent_item
                    
                    if c_path:
                        current_path = ""
                        for part in c_path.split('/'):
                            part = part.strip()
                            if not part: continue
                            current_path = f"{current_path}/{part}" if current_path else part
                            
                            if current_path not in root_items:
                                folder_item = QTreeWidgetItem([part])
                                folder_item.setIcon(0, qta.icon('fa5s.folder', color='#FFC107'))
                                current_parent.addChild(folder_item)
                                root_items[current_path] = folder_item
                            
                            current_parent = root_items[current_path]
                    
                    class_item = QTreeWidgetItem([c_name])
                    class_item.setIcon(0, qta.icon('fa5s.file-alt', color='#4CAF50'))
                    class_item.setData(0, Qt.UserRole, c_id)
                    current_parent.addChild(class_item)
                    
                self.sidebar.expandAll()
            
        except sqlite3.OperationalError as e:
            if "no such table" not in str(e).lower():
                print(f"Sidebar loading error: {e}")
            
        self.refresh_module_sidebar()

    def refresh_module_sidebar(self):
        self.module_sidebar.clear()
        settings = QSettings("MyCompany", "DatabaseManagerApp")
        db_path = settings.value("db_path", "")
        if not db_path or not os.path.exists(db_path): return
        
        try:
            with sqlite3.connect(db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT id, name, path FROM modules ORDER BY path ASC, name ASC")
                root_items = {}
                parent_item = self.module_sidebar.invisibleRootItem()
                
                for mid, mname, mpath in cur.fetchall():
                    mpath = (mpath or "").strip().strip('/')
                    current_parent = parent_item
                    
                    if mpath:
                        current_path = ""
                        for part in mpath.split('/'):
                            part = part.strip()
                            if not part: continue
                            current_path = f"{current_path}/{part}" if current_path else part
                            
                            if current_path not in root_items:
                                folder_item = QTreeWidgetItem([part])
                                folder_item.setIcon(0, qta.icon('fa5s.folder', color='#FFC107'))
                                current_parent.addChild(folder_item)
                                root_items[current_path] = folder_item
                            current_parent = root_items[current_path]
                    
                    item = QTreeWidgetItem([mname])
                    item.setIcon(0, qta.icon('fa5s.file-code', color='#2196F3'))
                    item.setData(0, Qt.UserRole, mid)
                    current_parent.addChild(item)
                    
                self.module_sidebar.expandAll()
        except sqlite3.OperationalError:
            pass

    def change_class(self, item, column):
        class_id = item.data(0, Qt.UserRole)
        if class_id:
            class_name = item.text(0)
            self.data_browser.load_table_data(class_id, class_name)

    def show_module_context_menu(self, position):
        item = self.module_sidebar.itemAt(position)
        if not item: return
        mid = item.data(0, Qt.UserRole)
        if not mid: return 
        
        menu = QMenu()
        act_run = menu.addAction(qta.icon('fa5s.play', color='green'), "Run Script")
        act_edit = menu.addAction(qta.icon('fa5s.edit'), "Edit Script in Builder")
        
        action = menu.exec(self.module_sidebar.viewport().mapToGlobal(position))
        if action == act_run:
            self.run_module_by_id(mid)
        elif action == act_edit:
            self.open_module_builder()

    def run_module_by_id(self, mid):
        settings = QSettings("MyCompany", "DatabaseManagerApp")
        db_path = settings.value("db_path", "")
        
        with sqlite3.connect(db_path) as conn:
            cur = conn.cursor()
            cur.execute("SELECT name, code FROM modules WHERE id = ?", (mid,))
            row = cur.fetchone()
        
        if not row or not row[1]:
            QMessageBox.warning(self, "Warning", "Module script is empty or missing.")
            return
            
        mname, code = row
        out_path = settings.value("module_output_path", os.path.join(os.path.expanduser("~"), "module_output.txt"))
        
        def local_get_objects(class_name):
            if pd is None: raise ImportError("Pandas library is not installed.")
            with sqlite3.connect(db_path) as conn:
                cur = conn.cursor()
                cur.execute("SELECT id FROM classes WHERE name = ?", (class_name,))
                r = cur.fetchone()
                if not r:
                    raise ValueError(f"Class '{class_name}' not found.")
                cls_id = r[0]
                
            view_name, safe_table_name = sync_physical_table(db_path, cls_id, class_name, parent_widget=self)
            if not view_name: raise RuntimeError(f"Failed to sync schema for '{class_name}'.")
                
            full_view_name = f"full_view_{safe_table_name}"
            with sqlite3.connect(db_path) as conn:
                df = pd.read_sql_query(f"SELECT * FROM {qid(full_view_name)}", conn)
            return df
            
        context = {'get_objects': local_get_objects, 'pd': pd}
        
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(f"--- Running Module: {mname} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n\n")
                with redirect_stdout(f), redirect_stderr(f): exec(code, context)
            QMessageBox.information(self, "Success", f"Script '{mname}' executed successfully.\nOutput saved to:\n{out_path}")
        except Exception as e:
            with open(out_path, 'a', encoding='utf-8') as f:
                f.write("\n\n--- RUNTIME ERROR ---\n")
                traceback.print_exc(file=f)
            QMessageBox.warning(self, "Script Error", f"An error occurred during execution of '{mname}'.\nCheck the output file for details:\n{out_path}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion") 
    app.setWindowIcon(get_app_icon()) 
    
    app_font = app.font()
    app_font.setPointSize(10)
    app.setFont(app_font)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())